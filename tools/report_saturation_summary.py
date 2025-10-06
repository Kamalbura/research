#!/usr/bin/env python3
"""Summarise saturation run artifacts for each suite.

This script inspects the JSON saturation summary emitted by the scheduler
along with the combined workbook to build a per-suite report. It can emit a
human-readable text summary or JSON suitable for further processing.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to parse the combined workbook") from exc


Numeric = Optional[float]


@dataclass
class RateSample:
    rate_mbps: float
    throughput_mbps: float
    loss_pct: float
    avg_rtt_ms: float
    min_rtt_ms: float
    max_rtt_ms: float


@dataclass
class SuiteReport:
    suite: str
    baseline_owd_p50_ms: Numeric = None
    baseline_owd_p95_ms: Numeric = None
    baseline_rtt_p50_ms: Numeric = None
    baseline_rtt_p95_ms: Numeric = None
    saturation_point_mbps: Numeric = None
    stop_cause: Optional[str] = None
    confidence: Numeric = None
    search_mode: Optional[str] = None
    resolution_mbps: Numeric = None
    rekey_ms: Numeric = None
    excel_path: Optional[str] = None
    rates: List[RateSample] = field(default_factory=list)
    telemetry: Dict[str, Dict[str, Any]] = field(default_factory=dict)

    def to_text(self) -> str:
        lines = [f"Suite: {self.suite}"]
        lines.append(
            "  Baseline OWD (p50/p95 ms): "
            f"{self._fmt_numeric(self.baseline_owd_p50_ms)} / "
            f"{self._fmt_numeric(self.baseline_owd_p95_ms)}"
        )
        lines.append(
            "  Baseline RTT (p50/p95 ms): "
            f"{self._fmt_numeric(self.baseline_rtt_p50_ms)} / "
            f"{self._fmt_numeric(self.baseline_rtt_p95_ms)}"
        )
        lines.append(f"  Saturation point (Mbps): {self._fmt_numeric(self.saturation_point_mbps)}")
        if self.stop_cause or self.confidence is not None:
            cause = self.stop_cause or "n/a"
            lines.append(
                f"  Stop cause: {cause} | confidence={self._fmt_numeric(self.confidence)}"
            )
        if self.search_mode or self.resolution_mbps is not None:
            mode = self.search_mode or "n/a"
            lines.append(
                f"  Search mode: {mode} | resolution={self._fmt_numeric(self.resolution_mbps)} Mbps"
            )
        lines.append(f"  Rekey duration (ms): {self._fmt_numeric(self.rekey_ms)}")
        if self.excel_path:
            lines.append(f"  Per-suite workbook: {self.excel_path}")
        if self.rates:
            lines.append("  Rates exercised:")
            for sample in sorted(self.rates, key=lambda s: s.rate_mbps):
                lines.append(
                    "    - "
                    f"{sample.rate_mbps:.1f} Mbps | thr={sample.throughput_mbps:.3f} Mbps | "
                    f"loss={sample.loss_pct:.3f}% | avg_rtt={sample.avg_rtt_ms:.3f} ms "
                    f"(min={sample.min_rtt_ms:.3f}, max={sample.max_rtt_ms:.3f})"
                )
        if self.telemetry:
            lines.append("  Telemetry summary:")
            for kind, stats in sorted(self.telemetry.items()):
                count = stats.get("count", 0)
                lines.append(f"    - {kind}: {count} samples")
                metrics = stats.get("metrics", {})
                for metric, values in sorted(metrics.items()):
                    lines.append(
                        "      "
                        f"{metric}: avg={self._fmt_numeric(values.get('avg'))} | "
                        f"min={self._fmt_numeric(values.get('min'))} | "
                        f"max={self._fmt_numeric(values.get('max'))}"
                    )
        return "\n".join(lines)

    @staticmethod
    def _fmt_numeric(value: Numeric) -> str:
        if value is None:
            return "n/a"
        return f"{value:.3f}"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "suite": self.suite,
            "baseline_owd_p50_ms": self.baseline_owd_p50_ms,
            "baseline_owd_p95_ms": self.baseline_owd_p95_ms,
            "baseline_rtt_p50_ms": self.baseline_rtt_p50_ms,
            "baseline_rtt_p95_ms": self.baseline_rtt_p95_ms,
            "saturation_point_mbps": self.saturation_point_mbps,
            "stop_cause": self.stop_cause,
            "confidence": self.confidence,
            "search_mode": self.search_mode,
            "resolution_mbps": self.resolution_mbps,
            "rekey_ms": self.rekey_ms,
            "excel_path": self.excel_path,
            "rates": [
                {
                    "rate_mbps": sample.rate_mbps,
                    "throughput_mbps": sample.throughput_mbps,
                    "loss_pct": sample.loss_pct,
                    "avg_rtt_ms": sample.avg_rtt_ms,
                    "min_rtt_ms": sample.min_rtt_ms,
                    "max_rtt_ms": sample.max_rtt_ms,
                }
                for sample in self.rates
            ],
            "telemetry": self.telemetry,
        }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract saturation run details from artifacts")
    parser.add_argument(
        "--summary-json",
        type=Path,
        default=None,
        help="Path to saturation summary JSON file",
    )
    parser.add_argument(
        "--combined-xlsx",
        type=Path,
        default=None,
        help="Path to combined workbook",
    )
    parser.add_argument(
        "--run-id",
        type=str,
        default=None,
        help="Run identifier (e.g. 1759766131) used to locate artifacts",
    )
    parser.add_argument(
        "--event-log",
        dest="event_logs",
        action="append",
        type=Path,
        help="Path to a JSON-lines control log (may be passed multiple times)",
    )
    parser.add_argument(
        "--format",
        choices=("text", "json"),
        default="text",
        help="Output format",
    )
    return parser.parse_args()


def load_json_summary(path: Path) -> Dict[str, Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    records: Dict[str, Dict[str, Any]] = {}
    for entry in data:
        suite = entry.get("suite")
        if not suite:
            continue
        records[suite] = {
            "baseline_owd_p50_ms": _coerce_float(entry.get("baseline_owd_p50_ms")),
            "baseline_owd_p95_ms": _coerce_float(entry.get("baseline_owd_p95_ms")),
            "baseline_rtt_p50_ms": _coerce_float(entry.get("baseline_rtt_p50_ms")),
            "baseline_rtt_p95_ms": _coerce_float(entry.get("baseline_rtt_p95_ms")),
            "saturation_point_mbps": _coerce_float(entry.get("saturation_point_mbps")),
            "stop_cause": entry.get("stop_cause"),
            "confidence": _coerce_float(entry.get("confidence")),
            "search_mode": entry.get("search_mode"),
            "resolution_mbps": _coerce_float(entry.get("resolution_mbps")),
            "rekey_ms": _coerce_float(entry.get("rekey_ms")),
            "excel_path": entry.get("excel_path"),
        }
    return records


def load_workbook_sheets(path: Path) -> Tuple[Dict[str, Any], Dict[str, List[dict]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    run_info = _load_run_info(workbook)
    sheets: Dict[str, List[dict]] = {}
    for name in ("gcs_summary", "saturation_overview", "saturation_samples", "telemetry_samples"):
        if name in workbook.sheetnames:
            sheets[name] = _sheet_as_dicts(workbook[name])
        else:
            sheets[name] = []
    workbook.close()
    return run_info, sheets


def _load_run_info(workbook) -> Dict[str, Any]:
    info: Dict[str, Any] = {}
    if "run_info" not in workbook.sheetnames:
        return info
    ws = workbook["run_info"]
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        key = row[0]
        if key is None:
            continue
        value = row[1] if len(row) > 1 else None
        info[str(key)] = value
    return info


def _sheet_as_dicts(ws) -> List[dict]:
    rows: List[dict] = []
    header: List[str] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            header = [str(col).strip() if col is not None else "" for col in row]
            continue
        if not header:
            continue
        record: Dict[str, Any] = {}
        for key, value in zip(header, row):
            if key:
                record[key] = value
        if record:
            rows.append(record)
    return rows


def build_suite_reports(
    json_records: Dict[str, Dict[str, Any]],
    sheets: Dict[str, List[dict]],
) -> Dict[str, SuiteReport]:
    reports: Dict[str, SuiteReport] = {}
    for suite, payload in json_records.items():
        reports[suite] = SuiteReport(
            suite=suite,
            baseline_owd_p50_ms=payload.get("baseline_owd_p50_ms"),
            baseline_owd_p95_ms=payload.get("baseline_owd_p95_ms"),
            baseline_rtt_p50_ms=payload.get("baseline_rtt_p50_ms"),
            baseline_rtt_p95_ms=payload.get("baseline_rtt_p95_ms"),
            saturation_point_mbps=payload.get("saturation_point_mbps"),
            stop_cause=payload.get("stop_cause"),
            confidence=payload.get("confidence"),
            search_mode=payload.get("search_mode"),
            resolution_mbps=payload.get("resolution_mbps"),
            rekey_ms=payload.get("rekey_ms"),
            excel_path=payload.get("excel_path"),
        )

    samples_by_suite: Dict[str, List[dict]] = defaultdict(list)
    for sample in sheets.get("saturation_samples", []):
        suite = sample.get("suite")
        if not suite:
            continue
        samples_by_suite[suite].append(sample)

    for suite, samples in samples_by_suite.items():
        report = reports.setdefault(suite, SuiteReport(suite=suite))
        for sample in samples:
            rate = _coerce_float(sample.get("rate_mbps"))
            thr = _coerce_float(sample.get("throughput_mbps"))
            loss = _coerce_float(sample.get("loss_pct"))
            avg_rtt = _coerce_float(sample.get("avg_rtt_ms"))
            min_rtt = _coerce_float(sample.get("min_rtt_ms"))
            max_rtt = _coerce_float(sample.get("max_rtt_ms"))
            if None in (rate, thr, loss, avg_rtt, min_rtt, max_rtt):
                continue
            report.rates.append(
                RateSample(
                    rate_mbps=rate,
                    throughput_mbps=thr,
                    loss_pct=loss,
                    avg_rtt_ms=avg_rtt,
                    min_rtt_ms=min_rtt,
                    max_rtt_ms=max_rtt,
                )
            )

    telemetry_samples = sheets.get("telemetry_samples", [])
    telemetry_stats = _summarise_telemetry(telemetry_samples)
    for suite, payload in telemetry_stats.items():
        report = reports.setdefault(suite, SuiteReport(suite=suite))
        report.telemetry = payload

    overview_by_suite = {row.get("suite"): row for row in sheets.get("saturation_overview", []) if row.get("suite")}
    for suite, row in overview_by_suite.items():
        report = reports.setdefault(suite, SuiteReport(suite=suite))
        if report.baseline_owd_p50_ms is None:
            report.baseline_owd_p50_ms = _coerce_float(row.get("baseline_owd_p50_ms"))
        if report.baseline_owd_p95_ms is None:
            report.baseline_owd_p95_ms = _coerce_float(row.get("baseline_owd_p95_ms"))
        if report.baseline_rtt_p50_ms is None:
            report.baseline_rtt_p50_ms = _coerce_float(row.get("baseline_rtt_p50_ms"))
        if report.baseline_rtt_p95_ms is None:
            report.baseline_rtt_p95_ms = _coerce_float(row.get("baseline_rtt_p95_ms"))
        if report.saturation_point_mbps is None:
            report.saturation_point_mbps = _coerce_float(row.get("saturation_point_mbps"))
        if report.stop_cause is None:
            report.stop_cause = row.get("stop_cause")
        if report.confidence is None:
            report.confidence = _coerce_float(row.get("confidence"))
        if report.search_mode is None:
            report.search_mode = row.get("search_mode")
        if report.resolution_mbps is None:
            report.resolution_mbps = _coerce_float(row.get("resolution_mbps"))
        if report.rekey_ms is None:
            report.rekey_ms = _coerce_float(row.get("rekey_ms"))
        if not report.excel_path and row.get("excel_path"):
            report.excel_path = str(row.get("excel_path"))

    return reports


def _summarise_telemetry(samples: Iterable[dict]) -> Dict[str, Dict[str, Dict[str, Any]]]:
    summary: Dict[str, Dict[str, Dict[str, Any]]] = defaultdict(lambda: defaultdict(lambda: {"count": 0, "metrics": {}}))
    for sample in samples:
        kind = sample.get("kind") or "unknown"
        suite = (
            sample.get("suite")
            or sample.get("new_suite")
            or sample.get("old_suite")
            or sample.get("current_suite")
            or "unknown"
        )
        bucket = summary[suite][kind]
        bucket["count"] = bucket.get("count", 0) + 1
        for key, value in sample.items():
            if key in {"kind", "session_id", "peer", "source"}:
                continue
            numeric = _coerce_float(value)
            if numeric is None:
                continue
            metrics = bucket.setdefault("metrics", {})
            entry = metrics.setdefault(key, {"sum": 0.0, "count": 0, "min": numeric, "max": numeric})
            entry["sum"] += numeric
            entry["count"] += 1
            entry["min"] = min(entry["min"], numeric)
            entry["max"] = max(entry["max"], numeric)
    for suite, kinds in summary.items():
        for kind, stats in kinds.items():
            metrics = stats.get("metrics", {})
            for key, values in metrics.items():
                count = values.get("count", 0)
                avg = None
                if count:
                    avg = values["sum"] / count
                values["avg"] = avg
                del values["sum"]
                del values["count"]
    return summary


def _coerce_float(value: Any) -> Numeric:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return None
    return None


def emit_text(
    run_info: Dict[str, Any],
    reports: Dict[str, SuiteReport],
    events: Optional[Dict[str, Any]] = None,
) -> str:
    session_id = run_info.get("session_id", "unknown")
    generated = run_info.get("generated_utc", "unknown")
    run_id = run_info.get("run_id", "unknown")
    lines = [
        f"Session: {session_id}",
        f"Generated (UTC): {generated}",
        f"Run ID: {run_id}",
        f"Suites discovered: {len(reports)}",
        "",
    ]
    if events:
        lines.append("Event Timeline:")
        handshakes = events.get("handshakes", [])
        if handshakes:
            lines.append("  Handshakes:")
            for entry in sorted(handshakes, key=lambda item: item.get("ts") or ""):
                lines.append(
                    "    - "
                    f"{entry['ts']} :: suite={entry.get('suite_id', 'n/a')} "
                    f"source={entry.get('source', 'unknown')}"
                )
        rekeys = events.get("rekeys", [])
        if rekeys:
            lines.append("  Rekeys:")
            for entry in sorted(rekeys, key=lambda item: item.get("started_ts", "")):
                duration = entry.get("duration_ms")
                if duration is None:
                    duration_fmt = "n/a"
                else:
                    duration_fmt = f"{duration:.1f} ms"
                lines.append(
                    "    - "
                    f"{entry.get('started_ts', 'n/a')} → {entry.get('completed_ts', 'n/a')} | "
                    f"suite={entry.get('suite_id', 'n/a')} | rid={entry.get('rid', 'n/a')} | "
                    f"duration={duration_fmt} | source={entry.get('source', 'unknown')}"
                )
        warnings = events.get("warnings", [])
        if warnings:
            lines.append("  Warnings:")
            for entry in sorted(warnings, key=lambda item: item.get("ts") or ""):
                lines.append(
                    "    - "
                    f"{entry['ts']} :: {entry.get('msg', 'warning')} (source={entry.get('source', 'unknown')})"
                )
        lines.append("")
    for suite in sorted(reports):
        report = reports[suite]
        lines.append(report.to_text())
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def emit_json(
    run_info: Dict[str, Any],
    reports: Dict[str, SuiteReport],
    events: Optional[Dict[str, Any]] = None,
) -> str:
    payload = {
        "session_id": run_info.get("session_id"),
        "generated_utc": run_info.get("generated_utc"),
        "run_id": run_info.get("run_id"),
        "suites": [reports[name].to_dict() for name in sorted(reports)],
    }
    if events is not None:
        payload["events"] = events
    return json.dumps(payload, indent=2) + "\n"


def _discover_artifacts(
    summary_path: Optional[Path],
    workbook_path: Optional[Path],
    run_id: Optional[str],
) -> Tuple[Path, Path, Optional[str]]:
    if summary_path and workbook_path:
        run_id = run_id or _extract_run_id(summary_path) or _extract_run_id(workbook_path)
        return summary_path, workbook_path, run_id

    if run_id:
        inferred_summary = Path("logs/auto/gcs") / f"saturation_summary_run_{run_id}.json"
        inferred_workbook = Path("output/gcs") / f"run_{run_id}" / f"run_{run_id}_combined.xlsx"
        if not summary_path:
            summary_path = inferred_summary
        if not workbook_path:
            workbook_path = inferred_workbook

    if summary_path and not workbook_path:
        inferred_run = _extract_run_id(summary_path)
        if inferred_run:
            candidate = Path("output/gcs") / f"run_{inferred_run}" / f"run_{inferred_run}_combined.xlsx"
            if candidate.exists():
                workbook_path = candidate
                run_id = run_id or inferred_run

    if workbook_path and not summary_path:
        inferred_run = _extract_run_id(workbook_path)
        if inferred_run:
            candidate = Path("logs/auto/gcs") / f"saturation_summary_run_{inferred_run}.json"
            if candidate.exists():
                summary_path = candidate
                run_id = run_id or inferred_run

    if not summary_path or not workbook_path:
        summary_path, workbook_path, run_id = _auto_discover_latest(run_id)

    if not summary_path.exists():
        raise SystemExit(f"Summary JSON not found: {summary_path}")
    if not workbook_path.exists():
        raise SystemExit(f"Combined workbook not found: {workbook_path}")
    return summary_path, workbook_path, run_id or _extract_run_id(summary_path)


def _auto_discover_latest(forced_run: Optional[str]) -> Tuple[Path, Path, Optional[str]]:
    logs_dir = Path("logs/auto/gcs")
    if forced_run:
        summary_candidate = logs_dir / f"saturation_summary_run_{forced_run}.json"
        workbook_candidate = Path("output/gcs") / f"run_{forced_run}" / f"run_{forced_run}_combined.xlsx"
        return summary_candidate, workbook_candidate, forced_run

    candidates: List[Tuple[str, Path, Path]] = []
    for summary_path in logs_dir.glob("saturation_summary_run_*.json"):
        run = _extract_run_id(summary_path)
        if not run:
            continue
        workbook_path = Path("output/gcs") / f"run_{run}" / f"run_{run}_combined.xlsx"
        candidates.append((run, summary_path, workbook_path))

    if not candidates:
        raise SystemExit("No saturation summary JSON files found under logs/auto/gcs")

    for run, summary_path, workbook_path in sorted(candidates, key=lambda item: item[0], reverse=True):
        if workbook_path.exists():
            return summary_path, workbook_path, run

    run, summary_path, workbook_path = max(candidates, key=lambda item: item[0])
    return summary_path, workbook_path, run


def _extract_run_id(path: Path) -> Optional[str]:
    match = re.search(r"run_(\d+)", str(path))
    if match:
        return match.group(1)
    return None


def summarise_event_logs(paths: Iterable[Path]) -> Dict[str, Any]:
    handshakes: List[Dict[str, Any]] = []
    rekeys: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    for path in paths:
        if not path.exists():
            continue
        _parse_event_log(path, handshakes, rekeys, warnings)
    rekeys.sort(key=lambda item: item.get("started_ts") or "")
    return {
        "handshakes": handshakes,
        "rekeys": rekeys,
        "warnings": warnings,
    }


def _parse_event_log(
    path: Path,
    handshakes: List[Dict[str, Any]],
    rekeys: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
) -> None:
    rid_state: Dict[str, Dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            text = line.strip()
            if not text or not text.startswith("{"):
                continue
            try:
                record = json.loads(text)
            except json.JSONDecodeError:
                continue
            ts = record.get("ts")
            msg = record.get("msg", "")
            level = record.get("level", "INFO")
            suite_id = record.get("suite_id")
            rid = record.get("rid")
            source = str(path)
            if msg == "PQC handshake completed successfully":
                handshakes.append({
                    "ts": ts,
                    "suite_id": suite_id,
                    "source": source,
                })
                continue
            if msg == "Control rekey negotiation started" and rid:
                rid_state[rid] = {
                    "ts": ts,
                    "suite_id": suite_id,
                    "source": source,
                }
                continue
            if msg == "Control rekey successful" and rid:
                started = rid_state.get(rid)
                started_ts = started.get("ts") if started else None
                completed_ts = ts
                duration_ms = None
                if started_ts and completed_ts:
                    duration_ms = _compute_duration_ms(started_ts, completed_ts)
                rekeys.append(
                    {
                        "rid": rid,
                        "suite_id": suite_id,
                        "started_ts": started_ts,
                        "completed_ts": completed_ts,
                        "duration_ms": duration_ms,
                        "source": source,
                    }
                )
                continue
            if level == "WARNING":
                warnings.append({
                    "ts": ts,
                    "msg": msg,
                    "source": source,
                })


def _compute_duration_ms(start_ts: str, end_ts: str) -> Optional[float]:
    start = _parse_iso_ts(start_ts)
    end = _parse_iso_ts(end_ts)
    if not start or not end:
        return None
    delta = end - start
    return delta.total_seconds() * 1000.0


def _parse_iso_ts(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def main() -> None:
    args = parse_args()
    summary_path, workbook_path, run_id = _discover_artifacts(
        args.summary_json, args.combined_xlsx, args.run_id
    )
    json_records = load_json_summary(summary_path)
    run_info, sheets = load_workbook_sheets(workbook_path)
    reports = build_suite_reports(json_records, sheets)
    if run_id and "run_id" not in run_info:
        run_info["run_id"] = run_id
    events = None
    if args.event_logs:
        events = summarise_event_logs(args.event_logs)
    if args.format == "json":
        output = emit_json(run_info, reports, events)
    else:
        output = emit_text(run_info, reports, events)

    results_dir = Path("results")
    results_dir.mkdir(parents=True, exist_ok=True)
    suffix = "json" if args.format == "json" else "txt"
    if run_id:
        report_path = results_dir / f"report_run_{run_id}.{suffix}"
    else:
        report_path = results_dir / f"report.{suffix}"
    report_path.write_text(output, encoding="utf-8")
    print(output, end="")
    print(f"[info] wrote {report_path}")


if __name__ == "__main__":
    main()
