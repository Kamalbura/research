#!/usr/bin/env python3
"""Drone-side scheduler that controls the GCS follower."""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import socket
import subprocess
import sys
import threading
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None

# Ensure project root is on sys.path so `import core` works when running this file
ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core import suites as suites_mod
from core.config import CONFIG


DRONE_HOST = CONFIG["DRONE_HOST"]
GCS_HOST = CONFIG["GCS_HOST"]

CONTROL_HOST = CONFIG.get("GCS_CONTROL_HOST") or GCS_HOST
CONTROL_PORT = int(CONFIG.get("GCS_CONTROL_PORT", CONFIG.get("DRONE_CONTROL_PORT", 48080)))

APP_SEND_HOST = CONFIG.get("DRONE_PLAINTEXT_HOST", "127.0.0.1")
APP_SEND_PORT = int(CONFIG.get("DRONE_PLAINTEXT_TX", 47003))
APP_RECV_HOST = CONFIG.get("DRONE_PLAINTEXT_HOST", "127.0.0.1")
APP_RECV_PORT = int(CONFIG.get("DRONE_PLAINTEXT_RX", 47004))

OUTDIR = Path("logs/auto/drone_scheduler")
SUITES_OUTDIR = OUTDIR / "suites"
SECRETS_DIR = Path("secrets/matrix")

PROXY_STATUS_PATH = OUTDIR / "drone_status.json"
PROXY_SUMMARY_PATH = OUTDIR / "drone_summary.json"
SUMMARY_CSV = OUTDIR / "summary.csv"
EVENTS_FILENAME = "scheduler_events.jsonl"

TELEMETRY_BIND_HOST = CONFIG.get("DRONE_TELEMETRY_BIND", "0.0.0.0")
TELEMETRY_PORT = int(
    CONFIG.get("DRONE_TELEMETRY_PORT")
    or CONFIG.get("GCS_TELEMETRY_PORT")
    or 52080
)

COMBINED_OUTPUT_DIR = Path(
    CONFIG.get("DRONE_COMBINED_OUTPUT_BASE")
    or os.getenv("DRONE_COMBINED_OUTPUT_BASE", "output/drone")
)


def ts() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def mkdirp(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def suite_outdir(suite: str) -> Path:
    target = SUITES_OUTDIR / suite
    mkdirp(target)
    return target


def resolve_suites(requested: Optional[Iterable[str]]) -> List[str]:
    available = list(suites_mod.list_suites())
    if not available:
        raise RuntimeError("No suites registered in core.suites; cannot proceed")
    if not requested:
        return available
    resolved: List[str] = []
    seen: Set[str] = set()
    for name in requested:
        info = suites_mod.get_suite(name)
        suite_id = info["suite_id"]
        if suite_id not in available:
            raise RuntimeError(f"Suite {name} not present in core registry")
        if suite_id not in seen:
            resolved.append(suite_id)
            seen.add(suite_id)
    return resolved


def preferred_initial_suite(candidates: List[str]) -> Optional[str]:
    configured = CONFIG.get("SIMPLE_INITIAL_SUITE")
    if not configured:
        return None
    try:
        suite_id = suites_mod.get_suite(configured)["suite_id"]
    except NotImplementedError:
        return None
    return suite_id if suite_id in candidates else None


def ctl_send(obj: dict, timeout: float = 2.0, retries: int = 4, backoff: float = 0.5) -> dict:
    last_exc: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            with socket.create_connection((CONTROL_HOST, CONTROL_PORT), timeout=timeout) as sock:
                sock.sendall((json.dumps(obj) + "\n").encode())
                sock.shutdown(socket.SHUT_WR)
                line = sock.makefile().readline()
                return json.loads(line.strip()) if line else {}
        except Exception as exc:
            last_exc = exc
            if attempt < retries:
                time.sleep(backoff * attempt)
                continue
            raise
    if last_exc:
        raise last_exc
    return {}


class Blaster:
    """Local UDP traffic generator with RTT sampling."""

    def __init__(
        self,
        send_host: str,
        send_port: int,
        recv_host: str,
        recv_port: int,
        events_path: Path,
        payload_bytes: int,
        sample_every: int,
    ) -> None:
        self.send_addr = (send_host, send_port)
        self.recv_addr = (recv_host, recv_port)
        self.payload_bytes = max(12, int(payload_bytes))
        self.sample_every = max(0, int(sample_every))
        self.tx = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        self.rx = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        self.rx.bind(self.recv_addr)
        self.rx.settimeout(0.001)
        try:
            sndbuf = int(os.getenv("DRONE_SOCK_SNDBUF", str(1 << 20)))
            rcvbuf = int(os.getenv("DRONE_SOCK_RCVBUF", str(1 << 20)))
            self.tx.setsockopt(socket.SOL_SOCKET, socket.SO_SNDBUF, sndbuf)
            self.rx.setsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF, rcvbuf)
        except Exception:
            pass
        mkdirp(events_path.parent)
        self.events = open(events_path, "w", encoding="utf-8")
        self.sent = 0
        self.rcvd = 0
        self.sent_bytes = 0
        self.rcvd_bytes = 0
        self.rtt_sum_ns = 0
        self.rtt_samples = 0
        self.rtt_max_ns = 0
        self.rtt_min_ns: Optional[int] = None
        self.pending: Dict[int, int] = {}

    def _log_event(self, payload: dict) -> None:
        self.events.write(json.dumps(payload) + "\n")

    def _maybe_log(self, kind: str, seq: int, t_ns: int) -> None:
        if self.sample_every == 0:
            return
        if kind == "send":
            if seq % self.sample_every:
                return
        else:
            if self.rcvd % self.sample_every:
                return
        self._log_event({"event": kind, "seq": seq, "t_ns": t_ns})

    def run(self, duration_s: float, rate_pps: int) -> None:
        stop_at = time.time() + max(0.0, duration_s)
        payload_pad = b"\x00" * (self.payload_bytes - 12)
        interval = 0.0 if rate_pps <= 0 else 1.0 / max(1, rate_pps)
        stop_event = threading.Event()
        rx_thread = threading.Thread(target=self._rx_loop, args=(stop_event,), daemon=True)
        rx_thread.start()
        seq = 0
        burst = 32 if interval == 0.0 else 1
        while time.time() < stop_at:
            sends_this_loop = burst
            while sends_this_loop > 0:
                if time.time() >= stop_at:
                    break
                t_send = time.time_ns()
                packet = seq.to_bytes(4, "big") + int(t_send).to_bytes(8, "big") + payload_pad
                try:
                    self.tx.sendto(packet, self.send_addr)
                    if self.sample_every == 0 or (self.sample_every and seq % self.sample_every == 0):
                        self.pending[seq] = int(t_send)
                    self.sent += 1
                    self.sent_bytes += len(packet)
                    self._maybe_log("send", seq, int(t_send))
                except Exception as exc:
                    self._log_event({"event": "send_error", "err": str(exc), "ts": ts()})
                seq += 1
                sends_this_loop -= 1
            if interval > 0.0:
                time.sleep(interval)
            elif (seq & 0x3FFF) == 0:
                time.sleep(0)
        tail_deadline = time.time() + 0.25
        while time.time() < tail_deadline:
            if not self._rx_once():
                time.sleep(0)
        stop_event.set()
        rx_thread.join(timeout=0.2)
        try:
            self.events.flush()
        except Exception:
            pass
        self.events.close()
        self.tx.close()
        self.rx.close()

    def _rx_loop(self, stop_event: threading.Event) -> None:
        while not stop_event.is_set():
            progressed = False
            for _ in range(32):
                if self._rx_once():
                    progressed = True
                else:
                    break
            if not progressed:
                time.sleep(0)

    def _rx_once(self) -> bool:
        try:
            data, _ = self.rx.recvfrom(65535)
        except socket.timeout:
            return False
        except Exception:
            return False
        t_recv = time.time_ns()
        self.rcvd += 1
        self.rcvd_bytes += len(data)
        if len(data) >= 12:
            seq = int.from_bytes(data[:4], "big")
            t_send = self.pending.pop(seq, None)
            if t_send is not None:
                rtt = t_recv - t_send
                self.rtt_sum_ns += rtt
                self.rtt_samples += 1
                if rtt > self.rtt_max_ns:
                    self.rtt_max_ns = rtt
                if self.rtt_min_ns is None or rtt < self.rtt_min_ns:
                    self.rtt_min_ns = rtt
                self._maybe_log("recv", seq, int(t_recv))
        return True


def read_json(path: Path) -> dict:
    try:
        with open(path, encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return {}


def read_local_proxy_status() -> dict:
    data = read_json(PROXY_STATUS_PATH)
    if data:
        return data
    return read_json(PROXY_SUMMARY_PATH)


def read_local_proxy_counters() -> dict:
    status = read_local_proxy_status()
    if isinstance(status, dict):
        counters = status.get("counters")
        if isinstance(counters, dict) and counters:
            return counters
        if any(key in status for key in ("enc_out", "enc_in", "rekeys_ok", "rekeys_fail")):
            return status
    return {}


def snapshot_local_proxy_artifacts(suite: str) -> None:
    target = suite_outdir(suite)
    try:
        if PROXY_STATUS_PATH.exists():
            shutil.copy(PROXY_STATUS_PATH, target / "drone_status.json")
        if PROXY_SUMMARY_PATH.exists():
            shutil.copy(PROXY_SUMMARY_PATH, target / "drone_summary.json")
    except Exception:
        pass


def start_drone_proxy(suite: str) -> tuple[subprocess.Popen, object]:
    suite_dir = SECRETS_DIR / suite
    pub = suite_dir / "gcs_signing.pub"
    if not pub.exists():
        raise FileNotFoundError(f"Missing GCS signing public key for suite {suite}: {pub}")

    OUTDIR.mkdir(parents=True, exist_ok=True)
    log_path = OUTDIR / f"drone_{time.strftime('%Y%m%d-%H%M%S')}.log"
    log_handle = open(log_path, "w", encoding="utf-8", errors="replace")

    os.environ["DRONE_HOST"] = DRONE_HOST
    os.environ["GCS_HOST"] = GCS_HOST
    os.environ["ENABLE_PACKET_TYPE"] = "1" if CONFIG.get("ENABLE_PACKET_TYPE", True) else "0"
    os.environ["STRICT_UDP_PEER_MATCH"] = "1" if CONFIG.get("STRICT_UDP_PEER_MATCH", True) else "0"

    proc = subprocess.Popen(
        [
            sys.executable,
            "-m",
            "core.run_proxy",
            "drone",
            "--suite",
            suite,
            "--peer-pubkey-file",
            str(pub),
            "--status-file",
            str(PROXY_STATUS_PATH),
            "--json-out",
            str(PROXY_SUMMARY_PATH),
        ],
        stdout=log_handle,
        stderr=subprocess.STDOUT,
        text=True,
    )
    return proc, log_handle


def wait_handshake(timeout: float = 20.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        status = read_local_proxy_status()
        state = status.get("state") if isinstance(status, dict) else None
        if state in {"running", "completed", "ready", "handshake_ok"}:
            return True
        time.sleep(0.3)
    return False


def read_remote_status() -> dict:
    status = ctl_send({"cmd": "status"}, timeout=1.5, retries=2)
    return status if isinstance(status, dict) else {}


def read_remote_counters() -> dict:
    status = read_remote_status()
    counters = status.get("counters") if isinstance(status, dict) else None
    return counters if isinstance(counters, dict) else {}


def wait_remote_active_suite(target: str, timeout: float = 10.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        status = read_remote_status()
        if status.get("suite") == target:
            return True
        time.sleep(0.2)
    return False


def wait_remote_rekey(target_suite: str, baseline: Dict[str, object], timeout: float = 20.0) -> str:
    start = time.time()
    baseline_ok = int(baseline.get("rekeys_ok", 0) or 0)
    baseline_fail = int(baseline.get("rekeys_fail", 0) or 0)
    while time.time() - start < timeout:
        status = read_remote_status()
        counters = status.get("counters") if isinstance(status, dict) else {}
        if not isinstance(counters, dict):
            time.sleep(0.4)
            continue
        rekeys_ok = int(counters.get("rekeys_ok", 0) or 0)
        rekeys_fail = int(counters.get("rekeys_fail", 0) or 0)
        last_suite = counters.get("last_rekey_suite") or status.get("suite") or ""
        if rekeys_fail > baseline_fail:
            return "fail"
        if rekeys_ok > baseline_ok and (not last_suite or last_suite == target_suite):
            return "ok"
        time.sleep(0.4)
    return "timeout"


def activate_suite(suite: str, is_first: bool) -> float:
    if is_first:
        try:
            ctl_send({"cmd": "mark", "suite": suite})
        except Exception:
            pass
        try:
            ctl_send({"cmd": "rekey_complete", "suite": suite, "status": "ok"})
        except Exception:
            pass
        wait_remote_active_suite(suite, timeout=5.0)
        return 0.0

    baseline = read_remote_counters()
    start_ns = time.time_ns()
    try:
        ctl_send({"cmd": "mark", "suite": suite})
    except Exception:
        pass

    rekey_status = "timeout"
    try:
        ctl_send({"cmd": "rekey", "suite": suite}, timeout=2.0)
        rekey_status = wait_remote_rekey(suite, baseline, timeout=15.0)
    except Exception as exc:
        rekey_status = f"error:{exc}"[:32]
    finally:
        try:
            ctl_send({"cmd": "rekey_complete", "suite": suite, "status": rekey_status})
        except Exception:
            pass
    wait_remote_active_suite(suite, timeout=5.0)
    return (time.time_ns() - start_ns) / 1_000_000


def run_suite(
    suite: str,
    is_first: bool,
    duration_s: float,
    payload_bytes: int,
    event_sample: int,
    pass_index: int,
    pre_gap: float,
    rate_pps: int,
) -> dict:
    rekey_ms = activate_suite(suite, is_first)

    events_path = suite_outdir(suite) / EVENTS_FILENAME
    start_mark_ns = time.time_ns() + int(max(pre_gap, 0.0) * 1e9) + int(0.150 * 1e9)
    try:
        ctl_send({"cmd": "schedule_mark", "suite": suite, "t0_ns": start_mark_ns})
    except Exception:
        pass

    print(
        f"[{ts()}] >>> START suite={suite} pass={pass_index} duration={duration_s:.1f}s rate={rate_pps}pps",
        flush=True,
    )
    if pre_gap > 0:
        time.sleep(pre_gap)

    blaster = Blaster(
        APP_SEND_HOST,
        APP_SEND_PORT,
        APP_RECV_HOST,
        APP_RECV_PORT,
        events_path,
        payload_bytes=payload_bytes,
        sample_every=event_sample,
    )
    start_perf_ns = time.perf_counter_ns()
    blaster.run(duration_s=duration_s, rate_pps=rate_pps)
    end_perf_ns = time.perf_counter_ns()

    remote_counters = read_remote_counters()
    local_counters = read_local_proxy_counters()
    snapshot_local_proxy_artifacts(suite)

    elapsed_s = max(1e-9, (end_perf_ns - start_perf_ns) / 1e9)
    pps = blaster.sent / elapsed_s
    throughput_mbps = (blaster.rcvd_bytes * 8) / (elapsed_s * 1_000_000)
    avg_rtt_ms = (blaster.rtt_sum_ns // max(1, blaster.rtt_samples)) / 1_000_000
    max_rtt_ms = blaster.rtt_max_ns / 1_000_000
    loss_pct = 0.0
    if blaster.sent:
        loss_pct = max(0.0, (blaster.sent - blaster.rcvd) * 100.0 / blaster.sent)

    row = {
        "pass": pass_index,
        "suite": suite,
        "duration_s": round(elapsed_s, 3),
        "sent": blaster.sent,
        "rcvd": blaster.rcvd,
        "pps": round(pps, 1),
        "throughput_mbps": round(throughput_mbps, 3),
        "rtt_avg_ms": round(avg_rtt_ms, 3),
        "rtt_max_ms": round(max_rtt_ms, 3),
        "rtt_samples": blaster.rtt_samples,
        "loss_pct": round(loss_pct, 3),
        "rekey_ms": round(rekey_ms, 3),
        "remote_enc_out": remote_counters.get("enc_out", 0),
        "remote_enc_in": remote_counters.get("enc_in", 0),
        "remote_rekeys_ok": remote_counters.get("rekeys_ok", 0),
        "remote_rekeys_fail": remote_counters.get("rekeys_fail", 0),
        "local_enc_out": local_counters.get("enc_out", 0),
        "local_enc_in": local_counters.get("enc_in", 0),
    }
    print(
        f"[{ts()}] <<< STOP suite={suite} sent={blaster.sent} rcvd={blaster.rcvd} loss={row['loss_pct']:.2f}% "
        f"thr={row['throughput_mbps']:.2f}Mb/s rtt_avg={row['rtt_avg_ms']:.3f}ms",
        flush=True,
    )
    return row


def write_summary(rows: List[dict]) -> None:
    if not rows:
        return
    mkdirp(OUTDIR)
    with open(SUMMARY_CSV, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"[{ts()}] wrote {SUMMARY_CSV}")


class TelemetryCollector:
    def __init__(self, host: str, port: int) -> None:
        self.host = host
        self.port = port
        self.stop_event = threading.Event()
        self.server: Optional[socket.socket] = None
        self.accept_thread: Optional[threading.Thread] = None
        self.client_threads: List[threading.Thread] = []
        self.samples: List[dict] = []
        self.lock = threading.Lock()
        self.enabled = True

    def start(self) -> None:
        try:
            srv = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            srv.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            srv.bind((self.host, self.port))
            srv.listen(8)
            srv.settimeout(0.5)
            self.server = srv
            self.accept_thread = threading.Thread(target=self._accept_loop, daemon=True)
            self.accept_thread.start()
            print(f"[{ts()}] telemetry collector on {self.host}:{self.port}")
        except Exception as exc:
            print(f"[WARN] telemetry collector disabled: {exc}")
            self.enabled = False
            if self.server:
                try:
                    self.server.close()
                except Exception:
                    pass
            self.server = None

    def _accept_loop(self) -> None:
        assert self.server is not None
        while not self.stop_event.is_set():
            try:
                conn, addr = self.server.accept()
            except socket.timeout:
                continue
            except OSError:
                break
            except Exception as exc:
                if not self.stop_event.is_set():
                    print(f"[WARN] telemetry accept error: {exc}")
                continue
            thread = threading.Thread(target=self._client_loop, args=(conn, addr), daemon=True)
            thread.start()
            self.client_threads.append(thread)

    def _client_loop(self, conn: socket.socket, addr) -> None:
        peer = f"{addr[0]}:{addr[1]}"
        try:
            conn.settimeout(1.0)
            with conn, conn.makefile("r", encoding="utf-8") as reader:
                for line in reader:
                    if self.stop_event.is_set():
                        break
                    data = line.strip()
                    if not data:
                        continue
                    try:
                        payload = json.loads(data)
                    except json.JSONDecodeError:
                        continue
                    payload.setdefault("collector_ts_ns", time.time_ns())
                    payload.setdefault("source", "gcs-follower")
                    payload.setdefault("peer", peer)
                    with self.lock:
                        self.samples.append(payload)
        except Exception:
            pass

    def snapshot(self) -> List[dict]:
        with self.lock:
            return list(self.samples)

    def stop(self) -> None:
        self.stop_event.set()
        if self.server:
            try:
                self.server.close()
            except Exception:
                pass
        if self.accept_thread and self.accept_thread.is_alive():
            self.accept_thread.join(timeout=1.5)
        for thread in self.client_threads:
            if thread.is_alive():
                thread.join(timeout=1.0)


def export_combined_excel(
    session_id: str,
    summary_rows: List[dict],
    telemetry_samples: List[dict],
) -> Optional[Path]:
    if Workbook is None:
        print("[WARN] openpyxl not available; skipping workbook export")
        return None
    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = "run_info"
    info_sheet.append(["generated_utc", ts()])
    info_sheet.append(["session_id", session_id])

    if summary_rows:
        sheet = workbook.create_sheet("summary")
        headers = list(summary_rows[0].keys())
        sheet.append(headers)
        for row in summary_rows:
            sheet.append([row.get(header, "") for header in headers])

    if telemetry_samples:
        sheet = workbook.create_sheet("telemetry")
        headers: List[str] = []
        for sample in telemetry_samples:
            for key in sample.keys():
                if key not in headers:
                    headers.append(key)
        sheet.append(headers)
        for sample in telemetry_samples:
            sheet.append([sample.get(key, "") for key in headers])

    COMBINED_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    target_path = COMBINED_OUTPUT_DIR / f"{session_id}_combined.xlsx"
    workbook.save(target_path)
    return target_path


def main() -> None:
    OUTDIR.mkdir(parents=True, exist_ok=True)
    SUITES_OUTDIR.mkdir(parents=True, exist_ok=True)

    parser = argparse.ArgumentParser(description="Drone scheduler controlling the GCS follower")
    parser.add_argument(
        "--traffic",
        choices=["blast"],
        default="blast",
        help="Traffic pattern (only blast supported)",
    )
    parser.add_argument(
        "--pre-gap",
        type=float,
        default=1.0,
        help="Seconds to wait after (re)key before sending",
    )
    parser.add_argument(
        "--inter-gap",
        type=float,
        default=15.0,
        help="Seconds to wait between suites",
    )
    parser.add_argument(
        "--duration",
        type=float,
        default=45.0,
        help="Active send window per suite",
    )
    parser.add_argument(
        "--rate",
        type=int,
        default=0,
        help="Packets/sec for blast; 0 = as fast as possible",
    )
    parser.add_argument(
        "--payload-bytes",
        type=int,
        default=256,
        help="UDP payload size in bytes",
    )
    parser.add_argument(
        "--event-sample",
        type=int,
        default=100,
        help="Log every Nth send/recv event (0 = disable)",
    )
    parser.add_argument(
        "--passes",
        type=int,
        default=1,
        help="Number of full sweeps across suites",
    )
    parser.add_argument("--suites", nargs="*", help="Optional subset of suites to exercise")
    parser.add_argument("--session-id", help="Identifier for output artifacts")
    parser.add_argument(
        "--no-local-proxy",
        action="store_true",
        help="Skip launching the local drone proxy (assumes external process)",
    )
    args = parser.parse_args()

    if args.duration <= 0:
        raise ValueError("--duration must be positive")
    if args.pre_gap < 0:
        raise ValueError("--pre-gap must be >= 0")
    if args.inter_gap < 0:
        raise ValueError("--inter-gap must be >= 0")
    if args.rate < 0:
        raise ValueError("--rate must be >= 0")
    if args.passes <= 0:
        raise ValueError("--passes must be >= 1")

    suites = resolve_suites(args.suites)
    if not suites:
        raise RuntimeError("No suites selected for execution")

    session_id = args.session_id or f"session_{int(time.time())}"

    initial_suite = preferred_initial_suite(suites)
    if initial_suite and suites[0] != initial_suite:
        suites = [initial_suite] + [s for s in suites if s != initial_suite]
        print(f"[{ts()}] reordered suites to start with {initial_suite}")

    telemetry_collector = TelemetryCollector(TELEMETRY_BIND_HOST, TELEMETRY_PORT)
    telemetry_collector.start()

    drone_proc: Optional[subprocess.Popen] = None
    drone_log = None

    try:
        if not args.no_local_proxy:
            drone_proc, drone_log = start_drone_proxy(suites[0])
            time.sleep(1.0)
            if drone_proc.poll() is not None:
                raise RuntimeError(f"drone proxy exited with {drone_proc.returncode}")

        reachable = False
        for attempt in range(6):
            try:
                resp = ctl_send({"cmd": "ping"}, timeout=1.0, retries=1)
                if resp.get("ok"):
                    reachable = True
                    break
            except Exception:
                pass
            time.sleep(0.5)
        if reachable:
            print(f"[{ts()}] follower reachable at {CONTROL_HOST}:{CONTROL_PORT}")
        else:
            print(f"[WARN] follower not reachable at {CONTROL_HOST}:{CONTROL_PORT}")

        if not wait_handshake(timeout=20.0):
            print(f"[WARN] local handshake not confirmed for {suites[0]}")

        summary_rows: List[dict] = []

        for pass_index in range(args.passes):
            for idx, suite in enumerate(suites):
                row = run_suite(
                    suite,
                    is_first=(pass_index == 0 and idx == 0),
                    duration_s=args.duration,
                    payload_bytes=args.payload_bytes,
                    event_sample=args.event_sample,
                    pass_index=pass_index,
                    pre_gap=args.pre_gap,
                    rate_pps=args.rate,
                )
                summary_rows.append(row)
                is_last_suite = idx == len(suites) - 1
                is_last_pass = pass_index == args.passes - 1
                if args.inter_gap > 0 and not (is_last_suite and is_last_pass):
                    time.sleep(args.inter_gap)

        write_summary(summary_rows)
        telemetry_samples: List[dict] = []
        if telemetry_collector.enabled:
            telemetry_samples = telemetry_collector.snapshot()
            telemetry_path = OUTDIR / f"telemetry_{session_id}.jsonl"
            with open(telemetry_path, "w", encoding="utf-8") as handle:
                for sample in telemetry_samples:
                    handle.write(json.dumps(sample) + "\n")
            print(f"[{ts()}] wrote {telemetry_path}")

        combined_path = export_combined_excel(session_id, summary_rows, telemetry_samples)
        if combined_path:
            print(f"[{ts()}] combined workbook written to {combined_path}")

    finally:
        try:
            ctl_send({"cmd": "stop"}, timeout=1.0, retries=1)
        except Exception:
            pass

        telemetry_collector.stop()

        if drone_proc:
            try:
                drone_proc.terminate()
                drone_proc.wait(timeout=5)
            except Exception:
                try:
                    drone_proc.kill()
                except Exception:
                    pass
        if drone_log:
            try:
                drone_log.close()
            except Exception:
                pass


if __name__ == "__main__":
    main()
