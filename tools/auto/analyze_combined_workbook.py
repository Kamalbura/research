#!/usr/bin/env python3
"""Summarise combined GCS workbook outputs.

This helper reads the ``*_combined.xlsx`` files produced by ``gcs_scheduler``
and emits a concise textual report. Optional charts can be generated when
``matplotlib`` is available. The script is intentionally lightweight so it can
run directly on the automation hosts without additional tooling.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - tooling dependency
    raise SystemExit("openpyxl is required to analyse combined workbooks") from exc


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT_ROOT = ROOT / "output" / "gcs"


@dataclass
class SuiteRecord:
    suite: str
    throughput_mbps: float
    loss_pct: float
    rekey_ms: float
    power_avg_w: float
    power_energy_j: float
    power_ok: bool
    rekeys_fail: int


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Summarise combined GCS workbook outputs")
    parser.add_argument(
        "--workbook",
        type=Path,
        help="Path to a *_combined.xlsx workbook. Defaults to the most recent file in output/gcs/",
    )
    parser.add_argument(
        "--charts",
        action="store_true",
        help="Generate PNG charts alongside the workbook (requires matplotlib)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Directory to write generated charts. Defaults to the workbook directory.",
    )
    return parser.parse_args(argv)


def find_latest_workbook(base_dir: Path) -> Optional[Path]:
    if not base_dir.exists():
        return None
    candidates = sorted(base_dir.rglob("*_combined.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def load_sheet_records(workbook, sheet_name: str) -> List[Dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: List[Dict[str, object]] = []
    for row in rows[1:]:
        record = {}
        for header, cell in zip(headers, row):
            if header:
                record[header] = cell
        if record:
            records.append(record)
    return records


def load_run_info(workbook) -> Dict[str, object]:
    if "run_info" not in workbook.sheetnames:
        return {}
    ws = workbook["run_info"]
    info: Dict[str, object] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = row[0]
        value = row[1] if len(row) > 1 else None
        if isinstance(key, str) and key:
            info[key] = value
    return info


def safe_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return 0.0


def as_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off", ""}:
        return False
    return False


def extract_suite_records(gcs_rows: Iterable[Dict[str, object]]) -> List[SuiteRecord]:
    records: List[SuiteRecord] = []
    for row in gcs_rows:
        suite = str(row.get("suite")) if row.get("suite") else "unknown"
        records.append(
            SuiteRecord(
                suite=suite,
                throughput_mbps=round(safe_float(row.get("throughput_mbps")), 3),
                loss_pct=round(safe_float(row.get("loss_pct")), 3),
                rekey_ms=round(safe_float(row.get("rekey_ms")), 3),
                power_avg_w=round(safe_float(row.get("power_avg_w")), 3),
                power_energy_j=round(safe_float(row.get("power_energy_j")), 3),
                power_ok=as_bool(row.get("power_capture_ok", False)),
                rekeys_fail=int(safe_float(row.get("rekeys_fail"))),
            )
        )
    return records


def summarise(records: Sequence[SuiteRecord]) -> Dict[str, object]:
    if not records:
        return {}
    best_throughput = max(records, key=lambda r: r.throughput_mbps)
    worst_loss = max(records, key=lambda r: r.loss_pct)
    slowest_rekey = max(records, key=lambda r: r.rekey_ms)
    total_energy = sum(r.power_energy_j for r in records)
    avg_power = 0.0
    power_samples = [r.power_avg_w for r in records if r.power_avg_w > 0]
    if power_samples:
        avg_power = sum(power_samples) / len(power_samples)
    return {
        "suite_count": len(records),
        "avg_throughput": sum(r.throughput_mbps for r in records) / len(records),
        "avg_loss": sum(r.loss_pct for r in records) / len(records),
        "best_throughput": (best_throughput.suite, best_throughput.throughput_mbps),
        "worst_loss": (worst_loss.suite, worst_loss.loss_pct),
        "slowest_rekey": (slowest_rekey.suite, slowest_rekey.rekey_ms),
        "total_energy": total_energy,
        "avg_power": avg_power,
        "power_gaps": [r.suite for r in records if not r.power_ok],
        "rekey_failures": [r.suite for r in records if r.rekeys_fail > 0],
    }


def print_report(workbook_path: Path, run_info: Dict[str, object], summary: Dict[str, object], records: Sequence[SuiteRecord]) -> None:
    print(f"Workbook: {workbook_path}")
    session_id = run_info.get("session_id") or run_info.get("Session")
    if session_id:
        print(f"Session: {session_id}")
    generated = run_info.get("generated_utc")
    if generated:
        print(f"Generated UTC: {generated}")
    print()

    if not records:
        print("No gcs_summary data found; nothing to report.")
        return

    if not summary:
        summary = summarise(records)

    header = (
        f"{'Suite':<34} | {'Thr Mb/s':>9} | {'Loss %':>7} | {'Rekey ms':>9} | "
        f"{'Power W':>8} | {'Energy J':>9} | {'Power':>7} | {'RekeyFail':>9}"
    )
    print(header)
    print("-" * len(header))
    for rec in records:
        power_flag = "OK" if rec.power_ok else "MISS"
        print(
            f"{rec.suite:<34} | {rec.throughput_mbps:>9.2f} | {rec.loss_pct:>7.2f} | {rec.rekey_ms:>9.2f} | "
            f"{rec.power_avg_w:>8.3f} | {rec.power_energy_j:>9.3f} | {power_flag:>7} | {rec.rekeys_fail:>9}"
        )

    print()
    print("Overall metrics:")
    print(f"  Suites analysed   : {summary['suite_count']}")
    print(f"  Avg throughput    : {summary['avg_throughput']:.2f} Mb/s")
    print(f"  Avg loss          : {summary['avg_loss']:.2f} %")
    best_suite, best_thr = summary['best_throughput']
    print(f"  Best throughput   : {best_suite} @ {best_thr:.2f} Mb/s")
    loss_suite, loss_val = summary['worst_loss']
    print(f"  Highest loss      : {loss_suite} @ {loss_val:.2f} %")
    rekey_suite, rekey_ms = summary['slowest_rekey']
    print(f"  Slowest rekey     : {rekey_suite} @ {rekey_ms:.2f} ms")
    print(f"  Total energy      : {summary['total_energy']:.3f} J")
    print(f"  Avg power (if any): {summary['avg_power']:.3f} W")
    if summary["power_gaps"]:
        print("  Missing power data:")
        for suite in summary["power_gaps"]:
            print(f"    - {suite}")
    if summary["rekey_failures"]:
        print("  Rekey failures    :")
        for suite in summary["rekey_failures"]:
            print(f"    - {suite}")


def maybe_generate_charts(records: Sequence[SuiteRecord], workbook_path: Path, output_dir: Optional[Path]) -> List[Path]:
    if not records:
        return []
    try:
        import matplotlib.pyplot as plt
    except ImportError:  # pragma: no cover - optional dependency
        print("[WARN] matplotlib not available; skipping chart generation", file=sys.stderr)
        return []

    out_dir = output_dir or workbook_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    suites = [r.suite for r in records]
    positions = list(range(len(suites)))

    chart_paths: List[Path] = []

    def save_chart(fig, name: str) -> None:
        path = out_dir / f"{workbook_path.stem}_{name}.png"
        fig.tight_layout()
        fig.savefig(path, dpi=160)
        chart_paths.append(path)
        plt.close(fig)

    # Throughput chart
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.bar(positions, [r.throughput_mbps for r in records], color="#1f77b4")
    ax.set_ylabel("Throughput (Mb/s)")
    ax.set_title("Per-suite throughput")
    ax.set_xticks(positions)
    ax.set_xticklabels(suites, rotation=60, ha="right")
    save_chart(fig, "throughput")

    # Loss chart
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.bar(positions, [r.loss_pct for r in records], color="#d62728")
    ax.set_ylabel("Loss (%)")
    ax.set_title("Per-suite packet loss")
    ax.set_xticks(positions)
    ax.set_xticklabels(suites, rotation=60, ha="right")
    save_chart(fig, "loss")

    # Rekey line chart
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.plot(positions, [r.rekey_ms for r in records], marker="o", color="#2ca02c")
    ax.set_ylabel("Rekey duration (ms)")
    ax.set_title("Rekey duration by suite")
    ax.set_xticks(positions)
    ax.set_xticklabels(suites, rotation=60, ha="right")
    save_chart(fig, "rekey")

    return chart_paths


def main(argv: Optional[Sequence[str]] = None) -> None:
    args = parse_args(argv)

    workbook_path: Optional[Path]
    if args.workbook:
        workbook_path = resolve_path(args.workbook)
        if not workbook_path.exists():
            raise SystemExit(f"Workbook not found: {workbook_path}")
    else:
        workbook_path = find_latest_workbook(DEFAULT_OUTPUT_ROOT)
        if not workbook_path:
            raise SystemExit(f"No *_combined.xlsx files found under {DEFAULT_OUTPUT_ROOT}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    run_info = load_run_info(workbook)
    gcs_rows = load_sheet_records(workbook, "gcs_summary")
    records = extract_suite_records(gcs_rows)
    summary = summarise(records)

    print_report(workbook_path, run_info, summary, records)

    if args.charts:
        chart_paths = maybe_generate_charts(records, workbook_path, args.output_dir)
        if chart_paths:
            print()
            print("Generated charts:")
            for path in chart_paths:
                print(f"  {path}")


def resolve_path(path: Path) -> Path:
    expanded = path.expanduser()
    return expanded if expanded.is_absolute() else (Path.cwd() / expanded)


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    main()
