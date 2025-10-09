#!/usr/bin/env python3
"""GCS scheduler that drives rekeys and UDP traffic using central configuration."""

from __future__ import annotations

import bisect
import csv
import errno
import io
import json
import math
import os
import socket
import struct
import subprocess
import sys
import threading
import time
from collections import deque, OrderedDict
from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, IO, Iterable, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:  # pragma: no cover
    Workbook = None
    BarChart = None
    LineChart = None
    Reference = None

def _ensure_core_importable() -> Path:
    root = Path(__file__).resolve().parents[2]
    root_str = str(root)
    if root_str not in sys.path:
        sys.path.insert(0, root_str)
    try:
        __import__("core")
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            f"Unable to import 'core'; repo root {root} missing from sys.path."
        ) from exc
    return root


ROOT = _ensure_core_importable()

from core import suites as suites_mod
from core.config import CONFIG
from tools.blackout_metrics import compute_blackout
from tools.merge_power import extract_power_fields
from tools.power_utils import calculate_transient_energy


DRONE_HOST = CONFIG["DRONE_HOST"]
GCS_HOST = CONFIG["GCS_HOST"]

CONTROL_PORT = int(CONFIG.get("DRONE_CONTROL_PORT", 48080))

APP_SEND_HOST = CONFIG.get("GCS_PLAINTEXT_HOST", "127.0.0.1")
APP_SEND_PORT = int(CONFIG.get("GCS_PLAINTEXT_TX", 47001))
APP_RECV_HOST = CONFIG.get("GCS_PLAINTEXT_HOST", "127.0.0.1")
APP_RECV_PORT = int(CONFIG.get("GCS_PLAINTEXT_RX", 47002))

OUTDIR = ROOT / "logs/auto/gcs"
SUITES_OUTDIR = OUTDIR / "suites"
SECRETS_DIR = ROOT / "secrets/matrix"

EXCEL_OUTPUT_DIR = ROOT / Path(
    CONFIG.get("GCS_EXCEL_OUTPUT")
    or os.getenv("GCS_EXCEL_OUTPUT", "output/gcs")
)

COMBINED_OUTPUT_DIR = ROOT / Path(
    CONFIG.get("GCS_COMBINED_OUTPUT_BASE")
    or os.getenv("GCS_COMBINED_OUTPUT_BASE", "output/gcs")
)

DRONE_MONITOR_BASE = ROOT / Path(
    CONFIG.get("DRONE_MONITOR_OUTPUT_BASE")
    or os.getenv("DRONE_MONITOR_OUTPUT_BASE", "output/drone")
)

TELEMETRY_BIND_HOST = CONFIG.get("GCS_TELEMETRY_BIND", "0.0.0.0")
TELEMETRY_PORT = int(
    CONFIG.get("GCS_TELEMETRY_PORT")
    or CONFIG.get("DRONE_TELEMETRY_PORT")
    or 52080
)

PROXY_STATUS_PATH = OUTDIR / "gcs_status.json"
PROXY_SUMMARY_PATH = OUTDIR / "gcs_summary.json"
SUMMARY_CSV = OUTDIR / "summary.csv"
EVENTS_FILENAME = "blaster_events.jsonl"
BLACKOUT_CSV = OUTDIR / "gcs_blackouts.csv"
STEP_RESULTS_PATH = OUTDIR / "step_results.jsonl"

SEQ_TS_OVERHEAD_BYTES = 12
UDP_HEADER_BYTES = 8
IPV4_HEADER_BYTES = 20
IPV6_HEADER_BYTES = 40
MIN_DELAY_SAMPLES = 30
HYSTERESIS_WINDOW = 3
MAX_BISECT_STEPS = 3
WARMUP_FRACTION = 0.1
MAX_WARMUP_SECONDS = 1.0
SATURATION_COARSE_RATES = [5, 25, 50, 75, 100, 125, 150, 175, 200]
SATURATION_LINEAR_RATES = [
    5,
    10,
    15,
    20,
    25,
    30,
    35,
    40,
    45,
    50,
    60,
    70,
    80,
    90,
    100,
    125,
    150,
    175,
    200,
]
SATURATION_SIGNALS = ("owd_p95_spike", "delivery_degraded", "loss_excess")
TELEMETRY_BUFFER_MAXLEN_DEFAULT = 100_000
REKEY_SETTLE_SECONDS = 1.5
CLOCK_OFFSET_THRESHOLD_NS = 50_000_000
CONSTANT_RATE_MBPS_DEFAULT = 8.0


def _compute_sampling_params(duration_s: float, event_sample: int, min_delay_samples: int) -> Tuple[int, int]:
    if event_sample <= 0:
        return 0, 0
    effective_sample = event_sample
    effective_min = max(0, min_delay_samples)
    if duration_s < 20.0:
        effective_sample = max(1, min(event_sample, 20))
        scale = max(duration_s, 5.0) / 20.0
        effective_min = max(10, int(math.ceil(effective_min * scale))) if effective_min else 0
    return effective_sample, effective_min


def _close_socket(sock: Optional[socket.socket]) -> None:
    if sock is None:
        return
    try:
        sock.close()
    except Exception:
        pass


def _close_file(handle: Optional[IO[str]]) -> None:
    if handle is None:
        return
    try:
        handle.flush()
    except Exception:
        pass
    try:
        handle.close()
    except Exception:
        pass


class P2Quantile:
    def __init__(self, p: float) -> None:
        if not 0.0 < p < 1.0:
            raise ValueError("p must be between 0 and 1")
        self.p = p
        self._initial: List[float] = []
        self._q: List[float] = []
        self._n: List[int] = []
        self._np: List[float] = []
        self._dn = [0.0, p / 2.0, p, (1.0 + p) / 2.0, 1.0]
        self.count = 0

    def add(self, sample: float) -> None:
        x = float(sample)
        self.count += 1
        if self.count <= 5:
            bisect.insort(self._initial, x)
            if self.count == 5:
                self._q = list(self._initial)
                self._n = [1, 2, 3, 4, 5]
                self._np = [1.0, 1.0 + 2.0 * self.p, 1.0 + 4.0 * self.p, 3.0 + 2.0 * self.p, 5.0]
            return

        if not self._q:
            # Should not happen, but guard for consistency
            self._q = list(self._initial)
            self._n = [1, 2, 3, 4, 5]
            self._np = [1.0, 1.0 + 2.0 * self.p, 1.0 + 4.0 * self.p, 3.0 + 2.0 * self.p, 5.0]

        if x < self._q[0]:
            self._q[0] = x
            k = 0
        elif x >= self._q[4]:
            self._q[4] = x
            k = 3
        else:
            k = 0
            for idx in range(4):
                if self._q[idx] <= x < self._q[idx + 1]:
                    k = idx
                    break

        for idx in range(k + 1, 5):
            self._n[idx] += 1

        for idx in range(5):
            self._np[idx] += self._dn[idx]

        for idx in range(1, 4):
            d = self._np[idx] - self._n[idx]
            if (d >= 1 and self._n[idx + 1] - self._n[idx] > 1) or (d <= -1 and self._n[idx - 1] - self._n[idx] < -1):
                step = 1 if d > 0 else -1
                candidate = self._parabolic(idx, step)
                if self._q[idx - 1] < candidate < self._q[idx + 1]:
                    self._q[idx] = candidate
                else:
                    self._q[idx] = self._linear(idx, step)
                self._n[idx] += step

    def value(self) -> float:
        if self.count == 0:
            return 0.0
        if self.count <= 5 and self._initial:
            rank = (self.count - 1) * self.p
            idx = max(0, min(len(self._initial) - 1, int(round(rank))))
            return float(self._initial[idx])
        if not self._q:
            return 0.0
        return float(self._q[2])

    def _parabolic(self, idx: int, step: int) -> float:
        numerator_left = self._n[idx] - self._n[idx - 1] + step
        numerator_right = self._n[idx + 1] - self._n[idx] - step
        denominator = self._n[idx + 1] - self._n[idx - 1]
        if denominator == 0:
            return self._q[idx]
        return self._q[idx] + (step / denominator) * (
            numerator_left * (self._q[idx + 1] - self._q[idx]) / max(self._n[idx + 1] - self._n[idx], 1)
            + numerator_right * (self._q[idx] - self._q[idx - 1]) / max(self._n[idx] - self._n[idx - 1], 1)
        )

    def _linear(self, idx: int, step: int) -> float:
        target = idx + step
        denominator = self._n[target] - self._n[idx]
        if denominator == 0:
            return self._q[idx]
        return self._q[idx] + step * (self._q[target] - self._q[idx]) / denominator


def wilson_interval(successes: int, n: int, z: float = 1.96) -> Tuple[float, float]:
    if n <= 0:
        return (0.0, 1.0)
    proportion = successes / n
    z2 = z * z
    denom = 1.0 + z2 / n
    center = (proportion + z2 / (2.0 * n)) / denom
    margin = (z * math.sqrt((proportion * (1.0 - proportion) / n) + (z2 / (4.0 * n * n)))) / denom
    return (max(0.0, center - margin), min(1.0, center + margin))


def ip_header_bytes_for_host(host: str) -> int:
    return IPV6_HEADER_BYTES if ":" in host else IPV4_HEADER_BYTES


APP_IP_HEADER_BYTES = ip_header_bytes_for_host(APP_SEND_HOST)


def ts() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def log_runtime_environment(component: str) -> None:
    preview = ";".join(sys.path[:5])
    print(f"[{ts()}] {component} python_exe={sys.executable}")
    print(f"[{ts()}] {component} cwd={Path.cwd()}")
    print(f"[{ts()}] {component} sys.path_prefix={preview}")


def _merge_defaults(defaults: dict, override: Optional[dict]) -> dict:
    result = deepcopy(defaults)
    if isinstance(override, dict):
        for key, value in override.items():
            if isinstance(value, dict) and isinstance(result.get(key), dict):
                merged = result[key].copy()
                merged.update(value)
                result[key] = merged
            else:
                result[key] = value
    return result


AUTO_GCS_DEFAULTS = {
    "session_prefix": "session",
    "traffic": "constant",
    "duration_s": 45.0,
    "pre_gap_s": 1.0,
    "inter_gap_s": 15.0,
    "payload_bytes": 256,
    "event_sample": 100,
    "passes": 1,
    "rate_pps": 0,
    "bandwidth_mbps": 0.0,
    "max_rate_mbps": 200.0,
    "sat_search": "auto",
    "sat_delivery_threshold": 0.85,
    "sat_loss_threshold_pct": 5.0,
    "sat_rtt_spike_factor": 1.6,
    "suites": None,
    "launch_proxy": True,
    "monitors_enabled": True,
    "telemetry_enabled": True,
    "telemetry_bind_host": TELEMETRY_BIND_HOST,
    "telemetry_port": TELEMETRY_PORT,
    "export_combined_excel": True,
    "power_capture": True,
}

AUTO_GCS_CONFIG = _merge_defaults(AUTO_GCS_DEFAULTS, CONFIG.get("AUTO_GCS"))

SATURATION_SEARCH_MODE = str(AUTO_GCS_CONFIG.get("sat_search") or "auto").lower()
SATURATION_RTT_SPIKE = float(AUTO_GCS_CONFIG.get("sat_rtt_spike_factor") or 1.6)
SATURATION_DELIVERY_THRESHOLD = float(AUTO_GCS_CONFIG.get("sat_delivery_threshold") or 0.85)
SATURATION_LOSS_THRESHOLD = float(AUTO_GCS_CONFIG.get("sat_loss_threshold_pct") or 5.0)


def mkdirp(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _atomic_write_bytes(path: Path, data: bytes, *, tmp_suffix: str = ".tmp", retries: int = 6, backoff: float = 0.05) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + tmp_suffix)
    fd = os.open(str(tmp_path), os.O_WRONLY | os.O_CREAT | os.O_TRUNC, 0o644)
    try:
        with os.fdopen(fd, "wb", closefd=True) as handle:
            handle.write(data)
            try:
                handle.flush()
                os.fsync(handle.fileno())
            except Exception:
                pass
    except Exception:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise

    delay = backoff
    last_exc: Optional[Exception] = None
    for attempt in range(retries):
        try:
            os.replace(tmp_path, path)
            return
        except PermissionError as exc:  # pragma: no cover - platform specific
            last_exc = exc
            if attempt == retries - 1:
                try:
                    os.remove(path)
                except FileNotFoundError:
                    pass
                except Exception:
                    pass
                try:
                    os.replace(tmp_path, path)
                    return
                except Exception as final_exc:
                    last_exc = final_exc
                    break
        except OSError as exc:  # pragma: no cover - platform specific
            if exc.errno not in (errno.EACCES, errno.EPERM):
                raise
            last_exc = exc
            if attempt == retries - 1:
                try:
                    os.remove(path)
                except FileNotFoundError:
                    pass
                except Exception:
                    pass
                try:
                    os.replace(tmp_path, path)
                    return
                except Exception as final_exc:
                    last_exc = final_exc
                    break
        time.sleep(delay)
        delay = min(delay * 2, 0.5)

    try:
        os.remove(tmp_path)
    except Exception:
        pass
    if last_exc is not None:
        raise last_exc


def _robust_copy(src: Path, dst: Path, attempts: int = 3, delay: float = 0.05) -> bool:
    for attempt in range(1, attempts + 1):
        try:
            data = src.read_bytes()
        except FileNotFoundError:
            return False
        except OSError as exc:
            print(f"[WARN] failed to read {src}: {exc}", file=sys.stderr)
            if attempt == attempts:
                return False
            time.sleep(delay)
            continue
        try:
            _atomic_write_bytes(dst, data)
            return True
        except Exception as exc:  # pragma: no cover - platform specific
            print(f"[WARN] failed to update {dst}: {exc}", file=sys.stderr)
            if attempt == attempts:
                return False
            time.sleep(delay)
    return False


def suite_outdir(suite: str) -> Path:
    target = SUITES_OUTDIR / suite
    mkdirp(target)
    return target


def _as_float(value: object) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(result):
        return None
    return result


def _rounded(value: object, digits: int) -> object:
    num = _as_float(value)
    if num is None:
        return ""
    return round(num, digits)


def _ns_to_ms(value: object) -> float:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return 0.0
    return round(num / 1_000_000.0, 3)


def _ns_to_us(value: object) -> float:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return 0.0
    return round(num / 1_000.0, 3)


def _flatten_handshake_metrics(metrics: Dict[str, object]) -> Dict[str, object]:
    base = {
        "handshake_role": "",
        "handshake_total_ms": 0.0,
        "handshake_wall_start_ns": 0,
        "handshake_wall_end_ns": 0,
        "handshake_kem_keygen_us": 0.0,
        "handshake_kem_encap_us": 0.0,
        "handshake_kem_decap_us": 0.0,
        "handshake_sig_sign_us": 0.0,
        "handshake_sig_verify_us": 0.0,
        "handshake_kdf_server_us": 0.0,
        "handshake_kdf_client_us": 0.0,
        "handshake_kem_pub_bytes": 0,
        "handshake_kem_ct_bytes": 0,
        "handshake_sig_bytes": 0,
        "handshake_auth_tag_bytes": 0,
        "handshake_shared_secret_bytes": 0,
        "handshake_server_hello_bytes": 0,
        "handshake_challenge_bytes": 0,
    }
    if not isinstance(metrics, dict) or not metrics:
        return base.copy()

    result = base.copy()

    def _as_int(value: object) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    result["handshake_role"] = str(metrics.get("role") or "")
    result["handshake_total_ms"] = _ns_to_ms(metrics.get("handshake_total_ns"))
    result["handshake_wall_start_ns"] = _as_int(metrics.get("handshake_wall_start_ns"))
    result["handshake_wall_end_ns"] = _as_int(metrics.get("handshake_wall_end_ns"))

    primitives = metrics.get("primitives") or {}
    if isinstance(primitives, dict):
        kem_metrics = primitives.get("kem") or {}
        if isinstance(kem_metrics, dict):
            result["handshake_kem_keygen_us"] = _ns_to_us(kem_metrics.get("keygen_ns"))
            result["handshake_kem_encap_us"] = _ns_to_us(kem_metrics.get("encap_ns"))
            result["handshake_kem_decap_us"] = _ns_to_us(kem_metrics.get("decap_ns"))
            result["handshake_kem_pub_bytes"] = _as_int(kem_metrics.get("public_key_bytes"))
            result["handshake_kem_ct_bytes"] = _as_int(kem_metrics.get("ciphertext_bytes"))
            result["handshake_shared_secret_bytes"] = _as_int(kem_metrics.get("shared_secret_bytes"))
        sig_metrics = primitives.get("signature") or {}
        if isinstance(sig_metrics, dict):
            result["handshake_sig_sign_us"] = _ns_to_us(sig_metrics.get("sign_ns"))
            result["handshake_sig_verify_us"] = _ns_to_us(sig_metrics.get("verify_ns"))
            if not result["handshake_sig_bytes"]:
                result["handshake_sig_bytes"] = _as_int(sig_metrics.get("signature_bytes"))

    result["handshake_kdf_server_us"] = _ns_to_us(metrics.get("kdf_server_ns"))
    result["handshake_kdf_client_us"] = _ns_to_us(metrics.get("kdf_client_ns"))

    artifacts = metrics.get("artifacts") or {}
    if isinstance(artifacts, dict):
        if not result["handshake_sig_bytes"]:
            result["handshake_sig_bytes"] = _as_int(artifacts.get("signature_bytes"))
        result["handshake_auth_tag_bytes"] = _as_int(artifacts.get("auth_tag_bytes"))
        result["handshake_server_hello_bytes"] = _as_int(artifacts.get("server_hello_bytes"))
        result["handshake_challenge_bytes"] = _as_int(artifacts.get("challenge_bytes"))

    return result


def resolve_suites(requested: Optional[Iterable[str]]) -> List[str]:
    suite_listing = suites_mod.list_suites()
    if isinstance(suite_listing, dict):
        available = list(suite_listing.keys())
    else:
        available = list(suite_listing)
    if not available:
        raise RuntimeError("No suites registered in core.suites; cannot proceed")

    if not requested:
        return available

    resolved: List[str] = []
    seen: Set[str] = set()
    for name in requested:
        info = suites_mod.get_suite(name)
        suite_id = info["suite_id"]
        if suite_id not in available:
            raise RuntimeError(f"Suite {name} not present in core registry")
        if suite_id not in seen:
            resolved.append(suite_id)
            seen.add(suite_id)
    return resolved


def preferred_initial_suite(candidates: List[str]) -> Optional[str]:
    configured = CONFIG.get("SIMPLE_INITIAL_SUITE")
    if not configured:
        return None
    try:
        suite_id = suites_mod.get_suite(configured)["suite_id"]
    except NotImplementedError:
        return None
    return suite_id if suite_id in candidates else None


def ctl_send(obj: dict, timeout: float = 2.0, retries: int = 4, backoff: float = 0.5) -> dict:
    last_exc: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            with socket.create_connection((DRONE_HOST, CONTROL_PORT), timeout=timeout) as sock:
                sock.sendall((json.dumps(obj) + "\n").encode())
                sock.shutdown(socket.SHUT_WR)
                line = sock.makefile().readline()
                return json.loads(line.strip()) if line else {}
        except Exception as exc:
            last_exc = exc
            if attempt < retries:
                time.sleep(backoff * attempt)
                continue
            raise
    if last_exc:
        raise last_exc
    return {}


def request_power_capture(suite: str, duration_s: float, start_ns: Optional[int]) -> dict:
    payload = {
        "cmd": "power_capture",
        "suite": suite,
        "duration_s": duration_s,
    }
    if start_ns is not None:
        payload["start_ns"] = int(start_ns)
    try:
        resp = ctl_send(payload, timeout=1.5, retries=2, backoff=0.4)
    except Exception as exc:
        print(f"[WARN] power_capture request failed: {exc}", file=sys.stderr)
        return {"ok": False, "error": str(exc)}
    return resp


def poll_power_status(max_wait_s: float = 12.0, poll_s: float = 0.6) -> dict:
    deadline = time.time() + max_wait_s
    last: dict = {}
    while time.time() < deadline:
        try:
            resp = ctl_send({"cmd": "power_status"}, timeout=1.5, retries=1, backoff=0.3)
        except Exception as exc:
            last = {"ok": False, "error": str(exc)}
            time.sleep(poll_s)
            continue
        last = resp
        if not resp.get("ok"):
            break
        if not resp.get("available", True):
            break
        if not resp.get("busy", False):
            break
        time.sleep(poll_s)
    return last


class Blaster:
    """High-rate UDP blaster with RTT sampling and throughput accounting."""

    def __init__(
        self,
        send_host: str,
        send_port: int,
        recv_host: str,
        recv_port: int,
        events_path: Optional[Path],
        payload_bytes: int,
        sample_every: int,
        offset_ns: int,
    ) -> None:
        self.payload_bytes = max(12, int(payload_bytes))
        self.sample_every = max(0, int(sample_every))
        self.offset_ns = offset_ns

        send_info = socket.getaddrinfo(send_host, send_port, 0, socket.SOCK_DGRAM)
        if not send_info:
            raise OSError(f"Unable to resolve send address {send_host}:{send_port}")
        send_family, _stype, _proto, _canon, send_sockaddr = send_info[0]

        recv_info = socket.getaddrinfo(recv_host, recv_port, send_family, socket.SOCK_DGRAM)
        if not recv_info:
            recv_info = socket.getaddrinfo(recv_host, recv_port, 0, socket.SOCK_DGRAM)
        if not recv_info:
            raise OSError(f"Unable to resolve recv address {recv_host}:{recv_port}")
        recv_family, _rstype, _rproto, _rcanon, recv_sockaddr = recv_info[0]

        self.tx = socket.socket(send_family, socket.SOCK_DGRAM)
        self.rx = socket.socket(recv_family, socket.SOCK_DGRAM)
        self.send_addr = send_sockaddr
        self.recv_addr = recv_sockaddr
        self.rx.bind(self.recv_addr)
        self.rx.settimeout(0.001)
        self.rx_burst = max(1, int(os.getenv("GCS_RX_BURST", "32")))
        self._lock = threading.Lock()
        self._run_active = threading.Event()
        self._rx_thread: Optional[threading.Thread] = None
        self._stop_event: Optional[threading.Event] = None
        self._closed = False
        try:
            # Allow overriding socket buffer sizes via environment variables
            # Use GCS_SOCK_SNDBUF and GCS_SOCK_RCVBUF if present, otherwise default to 1 MiB
            sndbuf = int(os.getenv("GCS_SOCK_SNDBUF", str(1 << 20)))
            rcvbuf = int(os.getenv("GCS_SOCK_RCVBUF", str(1 << 20)))
            self.tx.setsockopt(socket.SOL_SOCKET, socket.SO_SNDBUF, sndbuf)
            self.rx.setsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF, rcvbuf)
            actual_snd = self.tx.getsockopt(socket.SOL_SOCKET, socket.SO_SNDBUF)
            actual_rcv = self.rx.getsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF)
            print(
                f"[{ts()}] blaster UDP socket buffers: snd={actual_snd} rcv={actual_rcv}",
                flush=True,
            )
        except Exception:
            # best-effort; continue even if setting buffers fails
            pass

        family = self.tx.family if self.tx.family in (socket.AF_INET, socket.AF_INET6) else self.rx.family
        ip_bytes = IPV6_HEADER_BYTES if family == socket.AF_INET6 else IPV4_HEADER_BYTES
        self.wire_header_bytes = UDP_HEADER_BYTES + ip_bytes

        self.events_path = events_path
        self.events: Optional[IO[str]] = None
        if events_path is not None:
            mkdirp(events_path.parent)
            self.events = open(events_path, "w", encoding="utf-8")

        self.truncated = 0
        self.sent = 0
        self.rcvd = 0
        self.sent_bytes = 0
        self.rcvd_bytes = 0
        self.rtt_sum_ns = 0
        self.rtt_samples = 0
        self.rtt_max_ns = 0
        self.rtt_min_ns: Optional[int] = None
        self.pending: Dict[int, int] = {}
        self.rtt_p50 = P2Quantile(0.5)
        self.rtt_p95 = P2Quantile(0.95)
        self.owd_p50 = P2Quantile(0.5)
        self.owd_p95 = P2Quantile(0.95)
        self.owd_samples = 0
        self.owd_p50_ns = 0.0
        self.owd_p95_ns = 0.0
        self.rtt_p50_ns = 0.0
        self.rtt_p95_ns = 0.0

    def _log_event(self, payload: dict) -> None:
        # Buffered write; caller flushes at end of run()
        if self.events is None:
            return
        self.events.write(json.dumps(payload) + "\n")

    def _now(self) -> int:
        return time.time_ns() + self.offset_ns

    def _maybe_log(self, kind: str, seq: int, t_ns: int) -> None:
        if self.sample_every == 0:
            return
        if kind == "send":
            if seq % self.sample_every:
                return
        else:
            with self._lock:
                rcvd_count = self.rcvd
            if rcvd_count % self.sample_every:
                return
        self._log_event({"event": kind, "seq": seq, "t_ns": t_ns})

    def run(self, duration_s: float, rate_pps: int, max_packets: Optional[int] = None) -> None:
        if self._closed:
            raise RuntimeError("Blaster is closed")
        if self._run_active.is_set():
            raise RuntimeError("Blaster.run is already in progress")

        stop_at = self._now() + int(max(0.0, duration_s) * 1e9)
        payload_pad = b"\x00" * (self.payload_bytes - 12)
        interval_ns = 0.0 if rate_pps <= 0 else 1_000_000_000.0 / max(1, rate_pps)

        stop_event = threading.Event()
        self._stop_event = stop_event
        self._run_active.set()
        rx_thread = threading.Thread(target=self._rx_loop, args=(stop_event,), daemon=True)
        self._rx_thread = rx_thread
        rx_thread.start()

        with self._lock:
            self.pending.clear()

        seq = 0
        burst = 32 if interval_ns == 0.0 else 1
        next_send_ns = float(self._now())
        try:
            while self._now() < stop_at:
                if max_packets is not None:
                    with self._lock:
                        if self.sent >= max_packets:
                            break
                loop_progress = False
                sends_this_loop = burst
                while sends_this_loop > 0:
                    now_ns = self._now()
                    if now_ns >= stop_at:
                        break
                    if interval_ns > 0.0:
                        wait_ns = next_send_ns - now_ns
                        if wait_ns > 0:
                            time.sleep(min(wait_ns / 1_000_000_000.0, 0.001))
                            break
                    t_send = self._now()
                    packet = seq.to_bytes(4, "big") + int(t_send).to_bytes(8, "big") + payload_pad
                    try:
                        self.tx.sendto(packet, self.send_addr)
                    except Exception as exc:  # pragma: no cover - hard to surface in tests
                        self._log_event({"event": "send_error", "err": str(exc), "seq": seq, "ts": ts()})
                        break
                    t_send_int = int(t_send)
                    with self._lock:
                        if self.sample_every and (seq % self.sample_every == 0):
                            self.pending[seq] = t_send_int
                        self.sent += 1
                        self.sent_bytes += len(packet)
                    loop_progress = True
                    self._maybe_log("send", seq, t_send_int)
                    seq += 1
                    sends_this_loop -= 1
                    if interval_ns > 0.0:
                        next_send_ns = max(next_send_ns + interval_ns, float(t_send) + interval_ns)
                    if max_packets is not None:
                        with self._lock:
                            if self.sent >= max_packets:
                                break
                if interval_ns == 0.0 and (seq & 0x3FFF) == 0:
                    time.sleep(0)
                if not loop_progress:
                    time.sleep(0.0005)

            tail_deadline = self._now() + int(0.25 * 1e9)
            while self._now() < tail_deadline:
                time.sleep(0.0005)
        finally:
            stop_event.set()
            rx_thread.join(timeout=0.5)
            self._run_active.clear()
            self._rx_thread = None
            self._stop_event = None
            self.owd_p50_ns = self.owd_p50.value()
            self.owd_p95_ns = self.owd_p95.value()
            self.rtt_p50_ns = self.rtt_p50.value()
            self.rtt_p95_ns = self.rtt_p95.value()
            self._cleanup()
        _close_socket(self.tx)
        _close_socket(self.rx)

    def _rx_loop(self, stop_event: threading.Event) -> None:
        while not stop_event.is_set():
            if not self._run_active.is_set():
                break
            progressed = False
            for _ in range(self.rx_burst):
                if self._rx_once():
                    progressed = True
                else:
                    break
            if not progressed:
                time.sleep(0.0005)

    def _rx_once(self) -> bool:
        try:
            data, _ = self.rx.recvfrom(65535)
        except socket.timeout:
            return False
        except (socket.error, OSError) as exc:
            # Only log unexpected socket failures
            if not isinstance(exc, (ConnectionResetError, ConnectionRefusedError)):
                self._log_event({"event": "rx_error", "err": str(exc), "ts": ts()})
            return False

        t_recv = self._now()
        data_len = len(data)
        if data_len < 4:
            with self._lock:
                self.rcvd += 1
                self.rcvd_bytes += data_len
                self.truncated += 1
            return True

        seq = int.from_bytes(data[:4], "big")
        header_t_send = int.from_bytes(data[4:12], "big") if data_len >= 12 else None
        drone_recv_ns = int.from_bytes(data[-8:], "big") if data_len >= 20 else None

        log_recv = False
        with self._lock:
            self.rcvd += 1
            self.rcvd_bytes += data_len
            t_send = self.pending.pop(seq, None)
            if t_send is None:
                t_send = header_t_send

            if t_send is not None:
                rtt = t_recv - t_send
                if rtt >= 0:
                    self.rtt_sum_ns += rtt
                    self.rtt_samples += 1
                    if rtt > self.rtt_max_ns:
                        self.rtt_max_ns = rtt
                    if self.rtt_min_ns is None or rtt < self.rtt_min_ns:
                        self.rtt_min_ns = rtt
                    self.rtt_p50.add(rtt)
                    self.rtt_p95.add(rtt)
                    log_recv = True

            if t_send is not None and drone_recv_ns is not None:
                owd_up_ns = drone_recv_ns - t_send
                if 0 <= owd_up_ns <= 5_000_000_000:
                    self.owd_samples += 1
                    self.owd_p50.add(owd_up_ns)
                    self.owd_p95.add(owd_up_ns)
            if data_len < 20:
                self.truncated += 1

        if log_recv:
            self._maybe_log("recv", seq, int(t_recv))
        return True

    def _cleanup(self) -> None:
        if self.events:
            try:
                self.events.flush()
                self.events.close()
            except Exception:
                pass
            self.events = None


def wait_handshake(timeout: float = 20.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        if PROXY_STATUS_PATH.exists():
            try:
                with open(PROXY_STATUS_PATH, encoding="utf-8") as handle:
                    js = json.load(handle)
            except Exception:
                js = {}
            state = js.get("state") or js.get("status")
            if state in {"running", "completed", "ready", "handshake_ok"}:
                return True
        time.sleep(0.3)
    return False


def wait_active_suite(target: str, timeout: float = 10.0) -> bool:
    return wait_rekey_transition(target, timeout=timeout)


def wait_pending_suite(target: str, timeout: float = 18.0, stable_checks: int = 2) -> bool:
    deadline = time.time() + timeout
    stable = 0
    while time.time() < deadline:
        try:
            status = ctl_send({"cmd": "status"}, timeout=0.6, retries=1)
        except Exception:
            status = {}
        pending = status.get("pending_suite")
        suite = status.get("suite")
        if pending == target:
            stable += 1
            if stable >= stable_checks:
                return True
        elif suite == target and pending in (None, "", target):
            # Rekey may have already completed; treat as success.
            return True
        else:
            stable = 0
        time.sleep(0.2)
    return False


def wait_rekey_transition(target: str, timeout: float = 20.0, stable_checks: int = 3) -> bool:
    deadline = time.time() + timeout
    last_status: dict = {}
    stable = 0
    while time.time() < deadline:
        try:
            status = ctl_send({"cmd": "status"}, timeout=0.6, retries=1)
        except Exception:
            status = {}
        last_status = status
        suite = status.get("suite")
        pending = status.get("pending_suite")
        last_requested = status.get("last_requested_suite")
        if suite == target and (pending in (None, "", target)):
            stable += 1
            if stable >= stable_checks:
                if last_requested and last_requested not in (suite, target):
                    print(
                        f"[{ts()}] follower reports suite={suite} but last_requested={last_requested}; continuing anyway",
                        file=sys.stderr,
                    )
                return True
        else:
            stable = 0
        time.sleep(0.2)
    if last_status:
        print(
            f"[{ts()}] follower status before timeout: suite={last_status.get('suite')} pending={last_status.get('pending_suite')}",
            file=sys.stderr,
        )
    return False


def timesync() -> dict:
    t1 = time.time_ns()
    resp = ctl_send({"cmd": "timesync", "t1_ns": t1})
    t4 = time.time_ns()
    t2 = int(resp.get("t2_ns", t1))
    t3 = int(resp.get("t3_ns", t4))
    delay_ns = (t4 - t1) - (t3 - t2)
    offset_ns = ((t2 - t1) + (t3 - t4)) // 2
    return {"offset_ns": offset_ns, "rtt_ns": delay_ns}


def snapshot_proxy_artifacts(suite: str) -> None:
    target_dir = suite_outdir(suite)
    if PROXY_STATUS_PATH.exists():
        _robust_copy(PROXY_STATUS_PATH, target_dir / "gcs_status.json")
    if PROXY_SUMMARY_PATH.exists():
        _robust_copy(PROXY_SUMMARY_PATH, target_dir / "gcs_summary.json")


def start_gcs_proxy(initial_suite: str) -> tuple[subprocess.Popen, IO[str]]:
    key_path = SECRETS_DIR / initial_suite / "gcs_signing.key"
    if not key_path.exists():
        raise FileNotFoundError(f"Missing GCS signing key for suite {initial_suite}: {key_path}")

    mkdirp(OUTDIR)
    log_path = OUTDIR / f"gcs_{time.strftime('%Y%m%d-%H%M%S')}.log"
    log_handle: IO[str] = open(log_path, "w", encoding="utf-8", errors="replace")

    env = os.environ.copy()
    env["DRONE_HOST"] = DRONE_HOST
    env["GCS_HOST"] = GCS_HOST
    env["ENABLE_PACKET_TYPE"] = "1" if CONFIG.get("ENABLE_PACKET_TYPE", True) else "0"
    env["STRICT_UDP_PEER_MATCH"] = "1" if CONFIG.get("STRICT_UDP_PEER_MATCH", True) else "0"

    root_str = str(ROOT)
    existing_py_path = env.get("PYTHONPATH")
    if existing_py_path:
        if root_str not in existing_py_path.split(os.pathsep):
            env["PYTHONPATH"] = root_str + os.pathsep + existing_py_path
    else:
        env["PYTHONPATH"] = root_str

    proc = subprocess.Popen(
        [
            sys.executable,
            "-m",
            "core.run_proxy",
            "gcs",
            "--suite",
            initial_suite,
            "--gcs-secret-file",
            str(key_path),
            "--control-manual",
            "--status-file",
            str(PROXY_STATUS_PATH),
            "--json-out",
            str(PROXY_SUMMARY_PATH),
        ],
        stdin=subprocess.PIPE,
        stdout=log_handle,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
        env=env,
        cwd=str(ROOT),
    )
    return proc, log_handle


def read_proxy_stats_live() -> dict:
    try:
        with open(PROXY_STATUS_PATH, encoding="utf-8") as handle:
            js = json.load(handle)
    except Exception:
        return {}
    if isinstance(js, dict):
        counters = js.get("counters")
        if isinstance(counters, dict):
            return counters
        if any(k in js for k in ("enc_out", "enc_in")):
            return js
    return {}


def read_proxy_summary() -> dict:
    if not PROXY_SUMMARY_PATH.exists():
        return {}
    try:
        with open(PROXY_SUMMARY_PATH, encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return {}



def _read_proxy_counters() -> dict:

    counters = read_proxy_stats_live()

    if isinstance(counters, dict) and counters:

        return counters

    summary = read_proxy_summary()

    if isinstance(summary, dict):

        summary_counters = summary.get("counters")

        if isinstance(summary_counters, dict):

            return summary_counters

        if any(key in summary for key in ("enc_out", "enc_in", "rekeys_ok", "rekeys_fail", "last_rekey_suite")):

            return summary

    return {}





def wait_proxy_rekey(

    target_suite: str,

    baseline: Dict[str, object],

    *,

    timeout: float = 20.0,

    poll_interval: float = 0.4,

    proc: subprocess.Popen,

) -> str:

    start = time.time()

    baseline_ok = int(baseline.get("rekeys_ok", 0) or 0)

def _extract_companion_metrics(
    samples: List[dict],
    *,
    suite: str,
    start_ns: int,
    end_ns: int,
) -> Dict[str, object]:
    cpu_max = 0.0
    rss_max_bytes = 0
    pfc_sum = 0.0
    vh_sum = 0.0
    vv_sum = 0.0
    kin_count = 0

    for sample in samples:
        try:
            ts_ns = int(sample.get("timestamp_ns"))
        except (TypeError, ValueError):
            continue
        if ts_ns < start_ns or ts_ns > end_ns:
            continue
        sample_suite = str(sample.get("suite") or "").strip()
        if sample_suite and sample_suite != suite:
            continue

        kind = str(sample.get("kind") or "").lower()
        if kind == "system_sample":
            cpu_val = _as_float(sample.get("cpu_percent"))
            if cpu_val is not None:
                cpu_max = max(cpu_max, cpu_val)
            mem_mb = _as_float(sample.get("mem_used_mb"))
            if mem_mb is not None:
                rss_candidate = int(mem_mb * 1024 * 1024)
                rss_max_bytes = max(rss_max_bytes, rss_candidate)
        elif kind == "psutil_sample":
            cpu_val = _as_float(sample.get("cpu_percent"))
            if cpu_val is not None:
                cpu_max = max(cpu_max, cpu_val)
            rss_val = _as_float(sample.get("rss_bytes"))
            if rss_val is not None:
                rss_candidate = int(rss_val)
                rss_max_bytes = max(rss_max_bytes, rss_candidate)
        elif kind == "kinematics":
            pfc_val = _as_float(sample.get("predicted_flight_constraint_w"))
            vh_val = _as_float(sample.get("velocity_horizontal_mps"))
            vv_val = _as_float(sample.get("velocity_vertical_mps"))
            if pfc_val is not None:
                pfc_sum += pfc_val
            if vh_val is not None:
                vh_sum += vh_val
            if vv_val is not None:
                vv_sum += vv_val
            kin_count += 1

    avg_vh = vh_sum / kin_count if kin_count else 0.0
    avg_vv = vv_sum / kin_count if kin_count else 0.0
    avg_pfc = pfc_sum / kin_count if kin_count else 0.0

    return {
        "cpu_max_percent": round(cpu_max, 3),
        "max_rss_bytes": int(max(0, rss_max_bytes)),
        "pfc_watts": round(avg_pfc, 3),
        "kinematics_vh": round(avg_vh, 3),
        "kinematics_vv": round(avg_vv, 3),
    }


    baseline_fail = int(baseline.get("rekeys_fail", 0) or 0)

    while time.time() - start < timeout:

        if proc.poll() is not None:

            raise RuntimeError("GCS proxy exited during rekey")

        counters = _read_proxy_counters()

        if counters:

            rekeys_ok = int(counters.get("rekeys_ok", 0) or 0)

            rekeys_fail = int(counters.get("rekeys_fail", 0) or 0)

            last_suite = counters.get("last_rekey_suite") or counters.get("suite") or ""

            if rekeys_fail > baseline_fail:

                return "fail"

            if rekeys_ok > baseline_ok and (not last_suite or last_suite == target_suite):

                return "ok"

        time.sleep(poll_interval)

    return "timeout"


def activate_suite(
    gcs: subprocess.Popen,
    suite: str,
    is_first: bool,
) -> Tuple[float, Optional[int], Optional[int]]:

    if gcs.poll() is not None:

        raise RuntimeError("GCS proxy is not running; cannot continue")

    start_ns = time.time_ns()
    mark_ns: Optional[int] = None
    rekey_complete_ns: Optional[int] = None

    if is_first:
        mark_ns = None
        rekey_complete_ns = None

        if not wait_rekey_transition(suite, timeout=12.0):
            raise RuntimeError(f"Follower did not confirm initial suite {suite}")

    else:

        assert gcs.stdin is not None

        try:
            status_snapshot = ctl_send({"cmd": "status"}, timeout=0.6, retries=1)
        except Exception:
            status_snapshot = {}
        previous_suite = status_snapshot.get("suite")

        print(f"[{ts()}] rekey -> {suite}")

        gcs.stdin.write(suite + "\n")
        gcs.stdin.flush()

        baseline = _read_proxy_counters()

        mark_ns = time.time_ns()
        try:
            ctl_send({"cmd": "mark", "suite": suite, "kind": "rekey"})
        except Exception as exc:
            print(f"[WARN] control mark failed for {suite}: {exc}", file=sys.stderr)
        pending_ack = False
        pending_ack_error: Optional[str] = None
        try:
            pending_ack = wait_pending_suite(suite, timeout=12.0)
        except Exception as exc:
            pending_ack_error = str(exc)

        rekey_status = "timeout"

        try:

            result = wait_proxy_rekey(suite, baseline, timeout=24.0, proc=gcs)

            rekey_status = result

            if result == "timeout":

                print(f"[WARN] timed out waiting for proxy to activate suite {suite}", file=sys.stderr)

            elif result == "fail":

                print(f"[WARN] proxy reported failed rekey for suite {suite}", file=sys.stderr)

        except RuntimeError as exc:
            rekey_status = "error"
            raise
        except Exception as exc:
            rekey_status = "error"
            print(f"[WARN] error while waiting for proxy rekey {suite}: {exc}", file=sys.stderr)
        finally:
            try:
                rekey_complete_ns = time.time_ns()
                ctl_send({"cmd": "rekey_complete", "suite": suite, "status": rekey_status})
            except Exception as exc:
                print(f"[WARN] rekey_complete failed for {suite}: {exc}", file=sys.stderr)

        if rekey_status != "ok":
            if not pending_ack and pending_ack_error:
                print(
                    f"[WARN] follower pending status check failed for suite {suite}: {pending_ack_error}",
                    file=sys.stderr,
                )
            elif not pending_ack:
                print(
                    f"[WARN] follower did not acknowledge pending suite {suite} before proxy reported {rekey_status}",
                    file=sys.stderr,
                )
            if not previous_suite:
                raise RuntimeError(f"Proxy rekey to {suite} reported {rekey_status}; previous suite unknown")
            expected_suite = previous_suite
        else:
            expected_suite = suite

        if not wait_rekey_transition(expected_suite, timeout=24.0):
            raise RuntimeError(
                f"Follower did not confirm suite {expected_suite} after rekey status {rekey_status}"
            )

        if rekey_status != "ok":
            raise RuntimeError(f"Proxy reported rekey status {rekey_status} for suite {suite}")

    if REKEY_SETTLE_SECONDS > 0:
        time.sleep(REKEY_SETTLE_SECONDS)

    elapsed_ms = (time.time_ns() - start_ns) / 1_000_000
    return elapsed_ms, mark_ns, rekey_complete_ns




def run_suite(
    gcs: subprocess.Popen,
    suite: str,
    is_first: bool,
    duration_s: float,
    payload_bytes: int,
    event_sample: int,
    offset_ns: int,
    pass_index: int,
    traffic_mode: str,
    pre_gap: float,
    inter_gap_s: float,
    rate_pps: int,
    target_bandwidth_mbps: float,
    power_capture_enabled: bool,
    clock_offset_warmup_s: float,
    min_delay_samples: int,
) -> dict:
    rekey_duration_ms, rekey_mark_ns, rekey_complete_ns = activate_suite(gcs, suite, is_first)

    effective_sample_every, effective_min_delay = _compute_sampling_params(
        duration_s,
        event_sample,
        min_delay_samples,
    )

    events_path = suite_outdir(suite) / EVENTS_FILENAME
    start_mark_ns = time.time_ns() + offset_ns + int(0.150 * 1e9) + int(max(pre_gap, 0.0) * 1e9)
    try:
        ctl_send(
            {
                "cmd": "schedule_mark",
                "suite": suite,
                "t0_ns": start_mark_ns,
                "kind": "window",
            }
        )
    except Exception as exc:
        print(f"[WARN] schedule_mark failed for {suite}: {exc}", file=sys.stderr)

    power_request_ok = False
    power_request_error: Optional[str] = None
    power_status: dict = {}
    if power_capture_enabled:
        power_start_ns = time.time_ns() + offset_ns + int(max(pre_gap, 0.0) * 1e9)
        power_resp = request_power_capture(suite, duration_s, power_start_ns)
        power_request_ok = bool(power_resp.get("ok"))
        power_request_error = power_resp.get("error") if not power_request_ok else None
        if not power_request_ok and power_request_error:
            print(f"[WARN] power capture not scheduled: {power_request_error}", file=sys.stderr)
        banner = f"[{ts()}] ===== POWER: START in {pre_gap:.1f}s | suite={suite} | duration={duration_s:.1f}s mode={traffic_mode} ====="
    else:
        power_request_error = "disabled"
        banner = (
            f"[{ts()}] ===== TRAFFIC: START in {pre_gap:.1f}s | suite={suite} | duration={duration_s:.1f}s "
            f"mode={traffic_mode} (power capture disabled) ====="
        )
    print(banner)
    if pre_gap > 0:
        time.sleep(pre_gap)

    warmup_s = max(clock_offset_warmup_s, min(MAX_WARMUP_SECONDS, duration_s * WARMUP_FRACTION))
    start_wall_ns = time.time_ns()
    start_perf_ns = time.perf_counter_ns()
    sent_packets = 0
    rcvd_packets = 0
    rcvd_bytes = 0
    avg_rtt_ns = 0
    max_rtt_ns = 0
    rtt_samples = 0
    blaster_sent_bytes = 0

    wire_header_bytes = UDP_HEADER_BYTES + APP_IP_HEADER_BYTES

    if traffic_mode in {"blast", "constant"}:
        if warmup_s > 0:
            warmup_blaster = Blaster(
                APP_SEND_HOST,
                APP_SEND_PORT,
                APP_RECV_HOST,
                APP_RECV_PORT,
                events_path=None,
                payload_bytes=payload_bytes,
                sample_every=0,
                offset_ns=offset_ns,
            )
            warmup_blaster.run(duration_s=warmup_s, rate_pps=rate_pps)
        start_wall_ns = time.time_ns()
        start_perf_ns = time.perf_counter_ns()
        blaster = Blaster(
            APP_SEND_HOST,
            APP_SEND_PORT,
            APP_RECV_HOST,
            APP_RECV_PORT,
            events_path,
            payload_bytes=payload_bytes,
            sample_every=effective_sample_every if effective_sample_every > 0 else 0,
            offset_ns=offset_ns,
        )
        blaster.run(duration_s=duration_s, rate_pps=rate_pps)
        sent_packets = blaster.sent
        rcvd_packets = blaster.rcvd
        rcvd_bytes = blaster.rcvd_bytes
        blaster_sent_bytes = blaster.sent_bytes
        wire_header_bytes = getattr(blaster, "wire_header_bytes", wire_header_bytes)
        sample_count = max(1, blaster.rtt_samples)
        avg_rtt_ns = blaster.rtt_sum_ns // sample_count
        max_rtt_ns = blaster.rtt_max_ns
        rtt_samples = blaster.rtt_samples
    else:
        time.sleep(duration_s)

    end_wall_ns = time.time_ns()
    end_perf_ns = time.perf_counter_ns()
    if power_capture_enabled:
        print(f"[{ts()}] ===== POWER: STOP | suite={suite} =====")
    else:
        print(f"[{ts()}] ===== TRAFFIC: STOP | suite={suite} =====")

    snapshot_proxy_artifacts(suite)
    proxy_stats = read_proxy_stats_live() or read_proxy_summary()
    handshake_metrics_payload: Dict[str, object] = {}
    if isinstance(proxy_stats, dict):
        handshake_metrics_payload = proxy_stats.get("handshake_metrics") or {}
        if not isinstance(handshake_metrics_payload, dict):
            handshake_metrics_payload = {}
    handshake_fields = _flatten_handshake_metrics(handshake_metrics_payload)

    if power_capture_enabled and power_request_ok:
        power_status = poll_power_status(max_wait_s=max(6.0, duration_s * 0.25))
        if power_status.get("error"):
            print(f"[WARN] power status error: {power_status['error']}", file=sys.stderr)
        if power_status.get("busy"):
            power_note = "capture_incomplete:busy"
            if not power_status.get("error") and not power_request_error:
                power_status.setdefault("error", "capture_incomplete")

    power_summary = power_status.get("last_summary") if isinstance(power_status, dict) else None
    status_for_extract: Dict[str, Any] = {}
    if isinstance(power_status, dict) and power_status:
        status_for_extract = power_status
    elif power_summary:
        status_for_extract = {"last_summary": power_summary}
    power_fields = extract_power_fields(status_for_extract) if status_for_extract else {}
    power_capture_complete = bool(power_summary)
    power_error = None
    if not power_capture_complete:
        if isinstance(power_status, dict):
            power_error = power_status.get("error")
            if not power_error and power_status.get("busy"):
                power_error = "capture_incomplete"
        if power_error is None:
            power_error = power_request_error

    if not power_capture_enabled:
        power_note = "disabled"
    elif not power_request_ok:
        power_note = f"request_error:{power_error}" if power_error else "request_error"
    elif power_capture_complete:
        power_note = "ok"
    else:
        power_note = f"capture_incomplete:{power_error}" if power_error else "capture_incomplete"

    elapsed_s = max(1e-9, (end_perf_ns - start_perf_ns) / 1e9)
    pps = sent_packets / elapsed_s if elapsed_s > 0 else 0.0
    throughput_mbps = (rcvd_bytes * 8) / (elapsed_s * 1_000_000) if elapsed_s > 0 else 0.0
    sent_mbps = (blaster_sent_bytes * 8) / (elapsed_s * 1_000_000) if blaster_sent_bytes else 0.0
    delivered_ratio = throughput_mbps / sent_mbps if sent_mbps > 0 else 0.0
    avg_rtt_ms = avg_rtt_ns / 1_000_000
    max_rtt_ms = max_rtt_ns / 1_000_000

    app_packet_bytes = payload_bytes + SEQ_TS_OVERHEAD_BYTES
    wire_packet_bytes_est = app_packet_bytes + wire_header_bytes
    goodput_mbps = (rcvd_packets * payload_bytes * 8) / (elapsed_s * 1_000_000) if elapsed_s > 0 else 0.0
    wire_throughput_mbps_est = (
        (rcvd_packets * wire_packet_bytes_est * 8) / (elapsed_s * 1_000_000)
        if elapsed_s > 0
        else 0.0
    )
    if sent_mbps > 0:
        goodput_ratio = goodput_mbps / sent_mbps
        goodput_ratio = max(0.0, min(1.0, goodput_ratio))
    else:
        goodput_ratio = 0.0

    owd_p50_ms = 0.0
    owd_p95_ms = 0.0
    rtt_p50_ms = 0.0
    rtt_p95_ms = 0.0
    sample_quality = "disabled" if effective_sample_every == 0 else "low"
    owd_samples = 0

    if traffic_mode in {"blast", "constant"}:
        owd_p50_ms = blaster.owd_p50_ns / 1_000_000
        owd_p95_ms = blaster.owd_p95_ns / 1_000_000
        rtt_p50_ms = blaster.rtt_p50_ns / 1_000_000
        rtt_p95_ms = blaster.rtt_p95_ns / 1_000_000
        owd_samples = blaster.owd_samples
        if effective_sample_every > 0:
            if (
                effective_min_delay == 0
                or (blaster.rtt_samples >= effective_min_delay and blaster.owd_samples >= effective_min_delay)
            ):
                sample_quality = "ok"

    loss_pct = 0.0
    if sent_packets:
        loss_pct = max(0.0, (sent_packets - rcvd_packets) * 100.0 / sent_packets)
    loss_successes = max(0, sent_packets - rcvd_packets)
    loss_low, loss_high = wilson_interval(loss_successes, sent_packets)

    power_avg_w_val = power_fields.get("avg_power_w") if power_fields else None
    if power_avg_w_val is None and power_summary:
        power_avg_w_val = power_summary.get("avg_power_w")
    if power_avg_w_val is not None:
        try:
            power_avg_w_val = float(power_avg_w_val)
        except (TypeError, ValueError):
            power_avg_w_val = None
    power_energy_val = power_fields.get("energy_j") if power_fields else None
    if power_energy_val is None and power_summary:
        power_energy_val = power_summary.get("energy_j")
    if power_energy_val is not None:
        try:
            power_energy_val = float(power_energy_val)
        except (TypeError, ValueError):
            power_energy_val = None
    power_duration_val = power_fields.get("duration_s") if power_fields else None
    if power_duration_val is None and power_summary:
        power_duration_val = power_summary.get("duration_s")
    if power_duration_val is not None:
        try:
            power_duration_val = float(power_duration_val)
        except (TypeError, ValueError):
            power_duration_val = None
    power_summary_path_val = ""
    if power_fields and power_fields.get("summary_json_path"):
        power_summary_path_val = str(power_fields.get("summary_json_path") or "")
    elif power_summary:
        power_summary_path_val = str(power_summary.get("summary_json_path") or power_summary.get("csv_path") or "")
    power_csv_path_val = power_summary.get("csv_path") if power_summary else ""
    power_samples_val = power_summary.get("samples") if power_summary else 0
    power_avg_current_val = (
        round(power_summary.get("avg_current_a", 0.0), 6) if power_summary else 0.0
    )
    power_avg_voltage_val = (
        round(power_summary.get("avg_voltage_v", 0.0), 6) if power_summary else 0.0
    )
    power_sample_rate_val = (
        round(power_summary.get("sample_rate_hz", 0.0), 3) if power_summary else 0.0
    )

    companion_metrics = {
        "cpu_max_percent": 0.0,
        "max_rss_bytes": 0,
        "pfc_watts": 0.0,
        "kinematics_vh": 0.0,
        "kinematics_vv": 0.0,
    }
    if telemetry_collector and telemetry_collector.enabled:
        try:
            companion_metrics = _extract_companion_metrics(
                telemetry_collector.snapshot(),
                suite=suite,
                start_ns=start_wall_ns,
                end_ns=end_wall_ns,
            )
        except Exception as exc:
            print(f"[WARN] telemetry aggregation failed for suite {suite}: {exc}", file=sys.stderr)

    part_b_metrics = proxy_stats.get("part_b_metrics") if isinstance(proxy_stats.get("part_b_metrics"), dict) else None
    if not isinstance(part_b_metrics, dict):
        part_b_metrics = {
            key: proxy_stats.get(key)
            for key in (
                "kem_keygen_ms",
                "kem_encaps_ms",
                "kem_decap_ms",
                "sig_sign_ms",
                "sig_verify_ms",
                "primitive_total_ms",
                "pub_key_size_bytes",
                "ciphertext_size_bytes",
                "sig_size_bytes",
                "shared_secret_size_bytes",
            )
        }

    def _metric_ms(name: str) -> float:
        value = part_b_metrics.get(name)
        return _as_float(value) if value is not None else 0.0

    def _metric_int(name: str) -> int:
        value = part_b_metrics.get(name)
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    row = {
        "pass": pass_index,
        "suite": suite,
        "traffic_mode": traffic_mode,
        "pre_gap_s": round(pre_gap, 3),
        "inter_gap_s": round(inter_gap_s, 3),
        "duration_s": round(elapsed_s, 3),
        "sent": sent_packets,
        "rcvd": rcvd_packets,
        "pps": round(pps, 1),
        "target_rate_pps": rate_pps,
        "target_bandwidth_mbps": round(target_bandwidth_mbps, 3) if target_bandwidth_mbps else 0.0,
        "throughput_mbps": round(throughput_mbps, 3),
        "sent_mbps": round(sent_mbps, 3),
        "delivered_ratio": round(delivered_ratio, 3) if sent_mbps > 0 else 0.0,
        "goodput_mbps": round(goodput_mbps, 3),
        "wire_throughput_mbps_est": round(wire_throughput_mbps_est, 3),
        "app_packet_bytes": app_packet_bytes,
        "wire_packet_bytes_est": wire_packet_bytes_est,
        "cpu_max_percent": companion_metrics["cpu_max_percent"],
        "max_rss_bytes": companion_metrics["max_rss_bytes"],
        "pfc_watts": companion_metrics["pfc_watts"],
        "kinematics_vh": companion_metrics["kinematics_vh"],
        "kinematics_vv": companion_metrics["kinematics_vv"],
        "goodput_ratio": round(goodput_ratio, 3),
        "rtt_avg_ms": round(avg_rtt_ms, 3),
        "rtt_max_ms": round(max_rtt_ms, 3),
        "rtt_p50_ms": round(rtt_p50_ms, 3),
        "rtt_p95_ms": round(rtt_p95_ms, 3),
        "owd_p50_ms": round(owd_p50_ms, 3),
        "owd_p95_ms": round(owd_p95_ms, 3),
        "rtt_samples": rtt_samples,
        "owd_samples": owd_samples,
        "sample_every": effective_sample_every,
        "min_delay_samples": effective_min_delay,
        "sample_quality": sample_quality,
        "loss_pct": round(loss_pct, 3),
        "loss_pct_wilson_low": round(loss_low * 100.0, 3),
        "loss_pct_wilson_high": round(loss_high * 100.0, 3),
        "enc_out": proxy_stats.get("enc_out", 0),
        "enc_in": proxy_stats.get("enc_in", 0),
        "drops": proxy_stats.get("drops", 0),
        "rekeys_ok": proxy_stats.get("rekeys_ok", 0),
        "rekeys_fail": proxy_stats.get("rekeys_fail", 0),
        "start_ns": start_wall_ns,
        "end_ns": end_wall_ns,
        "scheduled_mark_ns": start_mark_ns,
        "rekey_mark_ns": rekey_mark_ns,
        "rekey_ok_ns": rekey_complete_ns,
        "rekey_ms": round(rekey_duration_ms, 3),
        "rekey_energy_mJ": 0.0,
        "power_request_ok": power_request_ok,
        "power_capture_ok": power_capture_complete,
        "power_note": power_note,
        "power_error": power_error,
        "power_avg_w": round(power_avg_w_val, 6) if power_avg_w_val is not None else 0.0,
        "power_energy_j": round(power_energy_val, 6) if power_energy_val is not None else 0.0,
        "power_samples": power_samples_val,
        "power_avg_current_a": power_avg_current_val,
        "power_avg_voltage_v": power_avg_voltage_val,
        "power_sample_rate_hz": power_sample_rate_val,
        "power_duration_s": round(power_duration_val, 3) if power_duration_val is not None else 0.0,
        "power_csv_path": power_csv_path_val or "",
        "power_summary_path": power_summary_path_val or "",
        "blackout_ms": None,
        "gap_max_ms": None,
        "gap_p99_ms": None,
        "steady_gap_ms": None,
        "recv_rate_kpps_before": None,
        "recv_rate_kpps_after": None,
        "proc_ns_p95": None,
        "pair_start_ns": None,
        "pair_end_ns": None,
        "blackout_error": None,
        "timing_guard_ms": None,
        "timing_guard_violation": False,
        "kem_keygen_ms": round(_metric_ms("kem_keygen_ms"), 6),
        "kem_encaps_ms": round(_metric_ms("kem_encaps_ms"), 6),
        "kem_decap_ms": round(_metric_ms("kem_decap_ms"), 6),
        "sig_sign_ms": round(_metric_ms("sig_sign_ms"), 6),
        "sig_verify_ms": round(_metric_ms("sig_verify_ms"), 6),
        "primitive_total_ms": round(_metric_ms("primitive_total_ms"), 6),
        "pub_key_size_bytes": _metric_int("pub_key_size_bytes"),
        "ciphertext_size_bytes": _metric_int("ciphertext_size_bytes"),
        "sig_size_bytes": _metric_int("sig_size_bytes"),
        "shared_secret_size_bytes": _metric_int("shared_secret_size_bytes"),
        "kem_keygen_mJ": 0.0,
        "kem_encaps_mJ": 0.0,
        "kem_decap_mJ": 0.0,
        "sig_sign_mJ": 0.0,
        "sig_verify_mJ": 0.0,
    }

    row.update(handshake_fields)

    row["handshake_energy_mJ"] = 0.0
    row["handshake_energy_error"] = ""
    if handshake_fields["handshake_wall_start_ns"] and handshake_fields["handshake_wall_end_ns"]:
        if isinstance(power_csv_path_val, str) and power_csv_path_val:
            try:
                energy_mj = calculate_transient_energy(
                    power_csv_path_val,
                    int(handshake_fields["handshake_wall_start_ns"]),
                    int(handshake_fields["handshake_wall_end_ns"]),
                )
                row["handshake_energy_mJ"] = round(energy_mj, 3)
            except (FileNotFoundError, ValueError) as exc:
                row["handshake_energy_error"] = str(exc)
            except Exception as exc:
                row["handshake_energy_error"] = str(exc)

    primitive_duration_map = {
        "kem_keygen_ms": row["kem_keygen_ms"],
        "kem_encaps_ms": row["kem_encaps_ms"],
        "kem_decap_ms": row["kem_decap_ms"],
        "sig_sign_ms": row["sig_sign_ms"],
        "sig_verify_ms": row["sig_verify_ms"],
    }
    duration_total_ms = sum(max(0.0, value) for value in primitive_duration_map.values())
    if duration_total_ms > 0 and row["handshake_energy_mJ"] > 0:
        for name, duration_ms in primitive_duration_map.items():
            if duration_ms <= 0:
                continue
            energy_key = name.replace("_ms", "_mJ")
            portion = duration_ms / duration_total_ms
            row[energy_key] = round(row["handshake_energy_mJ"] * portion, 3)

    rekey_energy_error: Optional[str] = None
    if (
        power_csv_path_val
        and isinstance(power_csv_path_val, str)
        and rekey_mark_ns is not None
        and rekey_complete_ns is not None
        and rekey_complete_ns > rekey_mark_ns
    ):
        try:
            energy_mj = calculate_transient_energy(
                power_csv_path_val,
                int(rekey_mark_ns),
                int(rekey_complete_ns),
            )
            row["rekey_energy_mJ"] = round(energy_mj, 3)
        except FileNotFoundError as exc:
            rekey_energy_error = str(exc)
        except ValueError as exc:
            rekey_energy_error = str(exc)
        except Exception as exc:
            rekey_energy_error = str(exc)

    if rekey_energy_error:
        row["rekey_energy_error"] = rekey_energy_error

    if power_summary:
        print(
            f"[{ts()}] power summary suite={suite} avg={power_summary.get('avg_power_w', 0.0):.3f} W "
            f"energy={power_summary.get('energy_j', 0.0):.3f} J samples={power_summary.get('samples', 0)}"
        )
    elif power_capture_enabled and power_request_ok and power_error:
        print(f"[{ts()}] power summary unavailable for suite={suite}: {power_error}")

    target_desc = f" target={target_bandwidth_mbps:.2f} Mb/s" if target_bandwidth_mbps > 0 else ""
    print(
        f"[{ts()}] <<< FINISH suite={suite} mode={traffic_mode} sent={sent_packets} rcvd={rcvd_packets} "
        f"pps~{pps:.0f} thr~{throughput_mbps:.2f} Mb/s sent~{sent_mbps:.2f} Mb/s loss={loss_pct:.2f}% "
        f"rtt_avg={avg_rtt_ms:.3f}ms rtt_max={max_rtt_ms:.3f}ms rekey={rekey_duration_ms:.2f}ms "
        f"enc_out={row['enc_out']} enc_in={row['enc_in']}{target_desc} >>>"
    )

    return row


def write_summary(rows: List[dict]) -> None:
    if not rows:
        return
    mkdirp(OUTDIR)
    headers = list(rows[0].keys())
    for attempt in range(3):
        try:
            buffer = io.StringIO()
            writer = csv.DictWriter(buffer, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
            _atomic_write_bytes(SUMMARY_CSV, buffer.getvalue().encode("utf-8"))
            print(f"[{ts()}] wrote {SUMMARY_CSV}")
            return
        except Exception as exc:
            if attempt == 2:
                print(f"[WARN] failed to write {SUMMARY_CSV}: {exc}", file=sys.stderr)
            time.sleep(0.1)


def _append_blackout_records(records: List[Dict[str, Any]]) -> None:
    if not records:
        return
    try:
        BLACKOUT_CSV.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "timestamp_utc",
            "session_id",
            "index",
            "pass",
            "suite",
            "traffic_mode",
            "rekey_mark_ns",
            "rekey_ok_ns",
            "scheduled_mark_ns",
            "blackout_ms",
            "gap_max_ms",
            "gap_p99_ms",
            "steady_gap_ms",
            "recv_rate_kpps_before",
            "recv_rate_kpps_after",
            "proc_ns_p95",
            "pair_start_ns",
            "pair_end_ns",
            "blackout_error",
        ]
        new_file = not BLACKOUT_CSV.exists()
        with BLACKOUT_CSV.open("a", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            if new_file:
                writer.writeheader()
            for record in records:
                writer.writerow(record)
        print(f"[{ts()}] updated {BLACKOUT_CSV} ({len(records)} rows)")
    except Exception as exc:
        print(f"[WARN] blackout log append failed: {exc}", file=sys.stderr)


def _append_step_results(payloads: List[Dict[str, Any]]) -> None:
    if not payloads:
        return
    try:
        STEP_RESULTS_PATH.parent.mkdir(parents=True, exist_ok=True)
        with STEP_RESULTS_PATH.open("a", encoding="utf-8") as handle:
            for payload in payloads:
                handle.write(json.dumps(payload) + "\n")
        print(f"[{ts()}] appended {len(payloads)} step records -> {STEP_RESULTS_PATH}")
    except Exception as exc:
        print(f"[WARN] step_results append failed: {exc}", file=sys.stderr)


def _enrich_summary_rows(
    rows: List[dict],
    *,
    session_id: str,
    drone_session_dir: Optional[Path],
    traffic_mode: str,
    pre_gap_s: float,
    duration_s: float,
    inter_gap_s: float,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    blackout_records: List[Dict[str, Any]] = []
    step_payloads: List[Dict[str, Any]] = []
    session_dir_exists = bool(drone_session_dir and drone_session_dir.exists())
    session_dir_str = str(drone_session_dir) if drone_session_dir else ""
    for index, row in enumerate(rows):
        mark_ns = row.get("rekey_mark_ns")
        ok_ns = row.get("rekey_ok_ns")
        metrics: Dict[str, Any] = {}
        blackout_error: Optional[str] = None
        if session_dir_exists and mark_ns and ok_ns and ok_ns >= mark_ns:
            try:
                metrics = compute_blackout(drone_session_dir, int(mark_ns), int(ok_ns))
            except Exception as exc:
                blackout_error = str(exc)
                metrics = {}
        else:
            if not session_dir_exists:
                blackout_error = "session_dir_unavailable"
            elif not mark_ns or not ok_ns:
                blackout_error = "missing_mark_or_ok"
            elif ok_ns is not None and mark_ns is not None and ok_ns < mark_ns:
                blackout_error = "invalid_timestamp_order"

        row["blackout_ms"] = metrics.get("blackout_ms")
        row["gap_max_ms"] = metrics.get("gap_max_ms")
        row["gap_p99_ms"] = metrics.get("gap_p99_ms")
        row["steady_gap_ms"] = metrics.get("steady_gap_ms")
        row["recv_rate_kpps_before"] = metrics.get("recv_rate_kpps_before")
        row["recv_rate_kpps_after"] = metrics.get("recv_rate_kpps_after")
        row["proc_ns_p95"] = metrics.get("proc_ns_p95")
        row["pair_start_ns"] = metrics.get("pair_start_ns")
        row["pair_end_ns"] = metrics.get("pair_end_ns")
        if blackout_error is None:
            blackout_error = metrics.get("error")
        row["blackout_error"] = blackout_error

        guard_ms = int(
            max(row.get("pre_gap_s", pre_gap_s) or 0.0, 0.0) * 1000.0
            + max(row.get("duration_s", duration_s) or 0.0, 0.0) * 1000.0
            + 10_000
        )
        row["timing_guard_ms"] = guard_ms
        rekey_ms = row.get("rekey_ms") or 0.0
        try:
            rekey_ms_val = float(rekey_ms)
        except (TypeError, ValueError):
            rekey_ms_val = 0.0
        timing_violation = bool(rekey_ms_val and rekey_ms_val > guard_ms)
        row["timing_guard_violation"] = timing_violation
        if timing_violation:
            print(
                f"[WARN] rekey duration {rekey_ms_val:.2f} ms exceeds guard {guard_ms} ms (suite={row.get('suite')} pass={row.get('pass')})",
                file=sys.stderr,
            )

        row.setdefault("traffic_mode", traffic_mode)
        row.setdefault("pre_gap_s", pre_gap_s)
        row.setdefault("inter_gap_s", inter_gap_s)

        blackout_records.append(
            {
                "timestamp_utc": ts(),
                "session_id": session_id,
                "index": index,
                "pass": row.get("pass"),
                "suite": row.get("suite"),
                "traffic_mode": row.get("traffic_mode"),
                "rekey_mark_ns": mark_ns or "",
                "rekey_ok_ns": ok_ns or "",
                "scheduled_mark_ns": row.get("scheduled_mark_ns") or "",
                "blackout_ms": row.get("blackout_ms"),
                "gap_max_ms": row.get("gap_max_ms"),
                "gap_p99_ms": row.get("gap_p99_ms"),
                "steady_gap_ms": row.get("steady_gap_ms"),
                "recv_rate_kpps_before": row.get("recv_rate_kpps_before"),
                "recv_rate_kpps_after": row.get("recv_rate_kpps_after"),
                "proc_ns_p95": row.get("proc_ns_p95"),
                "pair_start_ns": row.get("pair_start_ns"),
                "pair_end_ns": row.get("pair_end_ns"),
                "blackout_error": blackout_error or "",
            }
        )

        payload = dict(row)
        payload["ts_utc"] = ts()
        payload["session_id"] = session_id
        payload["session_dir"] = session_dir_str
        payload["index"] = index
        payload["blackout_error"] = blackout_error
        payload["timing_guard_ms"] = guard_ms
        payload["timing_guard_violation"] = timing_violation
        step_payloads.append(payload)

    return blackout_records, step_payloads


class SaturationTester:
    def __init__(
        self,
        suite: str,
        payload_bytes: int,
        duration_s: float,
        event_sample: int,
        offset_ns: int,
        output_dir: Path,
        max_rate_mbps: int,
        search_mode: str,
        delivery_threshold: float,
        loss_threshold: float,
        spike_factor: float,
        min_delay_samples: int,
    ) -> None:
        self.suite = suite
        self.payload_bytes = payload_bytes
        self.duration_s = duration_s
        self.event_sample = max(0, int(event_sample))
        self.offset_ns = offset_ns
        self.output_dir = output_dir
        self.max_rate_mbps = max_rate_mbps
        self.search_mode = search_mode
        self.delivery_threshold = delivery_threshold
        self.loss_threshold = loss_threshold
        self.spike_factor = spike_factor
        self.min_delay_samples = max(0, int(min_delay_samples))
        self.records: List[Dict[str, float]] = []
        self._rate_cache: Dict[int, Tuple[Dict[str, float], bool, Optional[str]]] = {}
        self._baseline: Optional[Dict[str, float]] = None
        self._signal_history = {key: deque(maxlen=HYSTERESIS_WINDOW) for key in SATURATION_SIGNALS}
        self._last_ok_rate: Optional[int] = None
        self._first_bad_rate: Optional[int] = None
        self._stop_cause: Optional[str] = None
        self._stop_samples = 0

    def run(self) -> Dict[str, Optional[float]]:
        self.records = []
        self._rate_cache.clear()
        self._baseline = None
        self._signal_history = {key: deque(maxlen=HYSTERESIS_WINDOW) for key in SATURATION_SIGNALS}
        self._last_ok_rate = None
        self._first_bad_rate = None
        self._stop_cause = None
        self._stop_samples = 0

        used_mode = self.search_mode
        if self.search_mode == "linear":
            self._linear_search()
        else:
            self._coarse_search()
            if self._first_bad_rate is not None and self._last_ok_rate is not None:
                self._bisect_search()
            elif self.search_mode == "bisect" and self._first_bad_rate is None:
                self._linear_search()
                used_mode = "linear"

        resolution = None
        if self._first_bad_rate is not None and self._last_ok_rate is not None:
            resolution = max(0, self._first_bad_rate - self._last_ok_rate)
        saturation_point = self._last_ok_rate if self._last_ok_rate is not None else self._first_bad_rate
        confidence = min(1.0, self._stop_samples / 200.0) if self._stop_samples > 0 else 0.0

        baseline = self._baseline or {}
        return {
            "suite": self.suite,
            "baseline_owd_p50_ms": baseline.get("owd_p50_ms"),
            "baseline_owd_p95_ms": baseline.get("owd_p95_ms"),
            "baseline_rtt_p50_ms": baseline.get("rtt_p50_ms"),
            "baseline_rtt_p95_ms": baseline.get("rtt_p95_ms"),
            "saturation_point_mbps": saturation_point,
            "stop_cause": self._stop_cause,
            "confidence": round(confidence, 3),
            "search_mode": used_mode,
            "resolution_mbps": resolution,
        }

    def _linear_search(self) -> None:
        for rate in SATURATION_LINEAR_RATES:
            if rate > self.max_rate_mbps:
                break
            _, is_bad, _ = self._evaluate_rate(rate)
            if is_bad:
                break

    def _coarse_search(self) -> None:
        for rate in SATURATION_COARSE_RATES:
            if rate > self.max_rate_mbps:
                break
            _, is_bad, _ = self._evaluate_rate(rate)
            if is_bad:
                break

    def _bisect_search(self) -> None:
        if self._first_bad_rate is None:
            return
        lo = self._last_ok_rate if self._last_ok_rate is not None else 0
        hi = self._first_bad_rate
        steps = 0
        while hi - lo > 5 and steps < MAX_BISECT_STEPS:
            mid = max(1, int(round((hi + lo) / 2)))
            if mid == hi or mid == lo:
                break
            _, is_bad, _ = self._evaluate_rate(mid)
            steps += 1
            metrics = self._rate_cache[mid][0]
            sample_ok = metrics.get("sample_quality") == "ok"
            if not sample_ok:
                is_bad = True
            if is_bad:
                if mid < hi:
                    hi = mid
                if self._first_bad_rate is None or mid < self._first_bad_rate:
                    self._first_bad_rate = mid
            else:
                if mid > lo:
                    lo = mid
                if self._last_ok_rate is None or mid > self._last_ok_rate:
                    self._last_ok_rate = mid

    def _evaluate_rate(self, rate: int) -> Tuple[Dict[str, float], bool, Optional[str]]:
        cached = self._rate_cache.get(rate)
        if cached:
            return cached

        metrics = self._run_rate(rate)
        metrics["suite"] = self.suite
        self.records.append(metrics)

        if self._baseline is None and metrics.get("sample_quality") == "ok":
            self._baseline = {
                "owd_p50_ms": metrics.get("owd_p50_ms"),
                "owd_p95_ms": metrics.get("owd_p95_ms"),
                "rtt_p50_ms": metrics.get("rtt_p50_ms"),
                "rtt_p95_ms": metrics.get("rtt_p95_ms"),
            }

        signals = self._classify_signals(metrics)
        is_bad = any(signals.values())
        cause = self._update_history(signals, rate, metrics)
        if is_bad:
            if self._first_bad_rate is None or rate < self._first_bad_rate:
                self._first_bad_rate = rate
        else:
            if metrics.get("sample_quality") == "ok":
                if self._last_ok_rate is None or rate > self._last_ok_rate:
                    self._last_ok_rate = rate

        result = (metrics, is_bad, cause)
        self._rate_cache[rate] = result
        return result

    def _classify_signals(self, metrics: Dict[str, float]) -> Dict[str, bool]:
        signals = {key: False for key in SATURATION_SIGNALS}
        baseline = self._baseline
        owd_spike = False
        if baseline:
            baseline_p95 = baseline.get("owd_p95_ms") or 0.0
            if baseline_p95 > 0:
                owd_p95 = metrics.get("owd_p95_ms", 0.0)
                owd_spike = owd_p95 >= baseline_p95 * self.spike_factor
        signals["owd_p95_spike"] = owd_spike

        goodput_ratio = metrics.get("goodput_ratio", 0.0)
        ratio_drop = goodput_ratio < self.delivery_threshold
        delivery_degraded = ratio_drop and owd_spike
        signals["delivery_degraded"] = delivery_degraded

        loss_flag = metrics.get("loss_pct", 0.0) > self.loss_threshold
        if metrics.get("sample_quality") != "ok" and loss_flag and not (delivery_degraded or owd_spike):
            loss_flag = False
        signals["loss_excess"] = loss_flag
        return signals

    def _update_history(
        self,
        signals: Dict[str, bool],
        rate: int,
        metrics: Dict[str, float],
    ) -> Optional[str]:
        cause = None
        for key in SATURATION_SIGNALS:
            history = self._signal_history[key]
            history.append(bool(signals.get(key)))
            if self._stop_cause is None and sum(history) >= 2:
                self._stop_cause = key
                self._stop_samples = max(metrics.get("rtt_samples", 0), metrics.get("owd_samples", 0))
                cause = key
        return cause

    def _run_rate(self, rate_mbps: int) -> Dict[str, float]:
        denominator = max(self.payload_bytes * 8, 1)
        rate_pps = int((rate_mbps * 1_000_000) / denominator)
        if rate_pps <= 0:
            rate_pps = 1
        events_path = self.output_dir / f"saturation_{rate_mbps}Mbps.jsonl"
        warmup_s = min(MAX_WARMUP_SECONDS, self.duration_s * WARMUP_FRACTION)
        effective_sample_every, effective_min_delay = _compute_sampling_params(
            self.duration_s,
            self.event_sample,
            self.min_delay_samples,
        )
        if warmup_s > 0:
            warmup_blaster = Blaster(
                APP_SEND_HOST,
                APP_SEND_PORT,
                APP_RECV_HOST,
                APP_RECV_PORT,
                events_path=None,
                payload_bytes=self.payload_bytes,
                sample_every=0,
                offset_ns=self.offset_ns,
            )
            warmup_blaster.run(duration_s=warmup_s, rate_pps=rate_pps)
        blaster = Blaster(
            APP_SEND_HOST,
            APP_SEND_PORT,
            APP_RECV_HOST,
            APP_RECV_PORT,
            events_path,
            payload_bytes=self.payload_bytes,
            sample_every=effective_sample_every if effective_sample_every > 0 else 0,
            offset_ns=self.offset_ns,
        )
        start = time.perf_counter()
        blaster.run(duration_s=self.duration_s, rate_pps=rate_pps)
        elapsed = max(1e-9, time.perf_counter() - start)

        sent_packets = blaster.sent
        rcvd_packets = blaster.rcvd
        sent_bytes = blaster.sent_bytes
        rcvd_bytes = blaster.rcvd_bytes

        pps_actual = sent_packets / elapsed if elapsed > 0 else 0.0
        throughput_mbps = (rcvd_bytes * 8) / (elapsed * 1_000_000) if elapsed > 0 else 0.0
        sent_mbps = (sent_bytes * 8) / (elapsed * 1_000_000) if sent_bytes else 0.0
        delivered_ratio = throughput_mbps / sent_mbps if sent_mbps > 0 else 0.0

        avg_rtt_ms = (blaster.rtt_sum_ns / max(1, blaster.rtt_samples)) / 1_000_000 if blaster.rtt_samples else 0.0
        min_rtt_ms = (blaster.rtt_min_ns or 0) / 1_000_000
        max_rtt_ms = blaster.rtt_max_ns / 1_000_000

        app_packet_bytes = self.payload_bytes + SEQ_TS_OVERHEAD_BYTES
        wire_header_bytes = getattr(blaster, "wire_header_bytes", UDP_HEADER_BYTES + APP_IP_HEADER_BYTES)
        wire_packet_bytes_est = app_packet_bytes + wire_header_bytes
        goodput_mbps = (
            (rcvd_packets * self.payload_bytes * 8) / (elapsed * 1_000_000)
            if elapsed > 0
            else 0.0
        )
        wire_throughput_mbps_est = (
            (rcvd_packets * wire_packet_bytes_est * 8) / (elapsed * 1_000_000)
            if elapsed > 0
            else 0.0
        )
        if sent_mbps > 0:
            goodput_ratio = goodput_mbps / sent_mbps
            goodput_ratio = max(0.0, min(1.0, goodput_ratio))
        else:
            goodput_ratio = 0.0

        loss_pct = 0.0
        if sent_packets:
            loss_pct = max(0.0, (sent_packets - rcvd_packets) * 100.0 / sent_packets)
        loss_low, loss_high = wilson_interval(max(0, sent_packets - rcvd_packets), sent_packets)

        sample_quality = "disabled" if effective_sample_every == 0 else "low"
        if effective_sample_every > 0:
            if (
                effective_min_delay == 0
                or (blaster.rtt_samples >= effective_min_delay and blaster.owd_samples >= effective_min_delay)
            ):
                sample_quality = "ok"
            if getattr(blaster, "truncated", 0) > 0:
                sample_quality = "low"

        return {
            "rate_mbps": float(rate_mbps),
            "pps": float(rate_pps),
            "pps_actual": round(pps_actual, 1),
            "sent_mbps": round(sent_mbps, 3),
            "throughput_mbps": round(throughput_mbps, 3),
            "goodput_mbps": round(goodput_mbps, 3),
            "wire_throughput_mbps_est": round(wire_throughput_mbps_est, 3),
            "goodput_ratio": round(goodput_ratio, 3),
            "loss_pct": round(loss_pct, 3),
            "loss_pct_wilson_low": round(loss_low * 100.0, 3),
            "loss_pct_wilson_high": round(loss_high * 100.0, 3),
            "delivered_ratio": round(delivered_ratio, 3) if sent_mbps > 0 else 0.0,
            "avg_rtt_ms": round(avg_rtt_ms, 3),
            "min_rtt_ms": round(min_rtt_ms, 3),
            "max_rtt_ms": round(max_rtt_ms, 3),
            "rtt_p50_ms": round(blaster.rtt_p50_ns / 1_000_000, 3),
            "rtt_p95_ms": round(blaster.rtt_p95_ns / 1_000_000, 3),
            "owd_p50_ms": round(blaster.owd_p50_ns / 1_000_000, 3),
            "owd_p95_ms": round(blaster.owd_p95_ns / 1_000_000, 3),
            "rtt_samples": blaster.rtt_samples,
            "owd_samples": blaster.owd_samples,
            "sample_every": effective_sample_every,
            "min_delay_samples": effective_min_delay,
            "sample_quality": sample_quality,
            "app_packet_bytes": app_packet_bytes,
            "wire_packet_bytes_est": wire_packet_bytes_est,
        }

    def export_excel(self, session_id: str, output_base: Path) -> Optional[Path]:
        if Workbook is None:
            print("[WARN] openpyxl not available; skipping Excel export")
            return None
        output_base.mkdir(parents=True, exist_ok=True)
        path = output_base / f"saturation_{self.suite}_{session_id}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Saturation"
        ws.append([
            "rate_mbps",
            "pps",
            "pps_actual",
            "sent_mbps",
            "throughput_mbps",
            "goodput_mbps",
            "wire_throughput_mbps_est",
            "goodput_ratio",
            "loss_pct",
            "loss_pct_wilson_low",
            "loss_pct_wilson_high",
            "delivered_ratio",
            "avg_rtt_ms",
            "min_rtt_ms",
            "max_rtt_ms",
            "rtt_p50_ms",
            "rtt_p95_ms",
            "owd_p50_ms",
            "owd_p95_ms",
            "rtt_samples",
            "owd_samples",
            "sample_quality",
            "app_packet_bytes",
            "wire_packet_bytes_est",
        ])
        for record in self.records:
            ws.append([
                record.get("rate_mbps", 0.0),
                record.get("pps", 0.0),
                record.get("pps_actual", 0.0),
                record.get("sent_mbps", 0.0),
                record.get("throughput_mbps", 0.0),
                record.get("goodput_mbps", 0.0),
                record.get("wire_throughput_mbps_est", 0.0),
                record.get("goodput_ratio", 0.0),
                record.get("loss_pct", 0.0),
                record.get("loss_pct_wilson_low", 0.0),
                record.get("loss_pct_wilson_high", 0.0),
                record.get("delivered_ratio", 0.0),
                record.get("avg_rtt_ms", 0.0),
                record.get("min_rtt_ms", 0.0),
                record.get("max_rtt_ms", 0.0),
                record.get("rtt_p50_ms", 0.0),
                record.get("rtt_p95_ms", 0.0),
                record.get("owd_p50_ms", 0.0),
                record.get("owd_p95_ms", 0.0),
                record.get("rtt_samples", 0),
                record.get("owd_samples", 0),
                record.get("sample_quality", "low"),
                record.get("app_packet_bytes", 0),
                record.get("wire_packet_bytes_est", 0),
            ])
        for attempt in range(3):
            try:
                buffer = io.BytesIO()
                wb.save(buffer)
                _atomic_write_bytes(path, buffer.getvalue())
                return path
            except OSError as exc:  # pragma: no cover - platform specific
                if attempt == 2:
                    print(f"[WARN] failed to save {path}: {exc}", file=sys.stderr)
            except Exception as exc:  # pragma: no cover - platform specific
                if attempt == 2:
                    print(f"[WARN] failed to write saturation workbook {path}: {exc}", file=sys.stderr)
            time.sleep(0.1)
        return None


class TelemetryCollector:
    def __init__(self, host: str, port: int) -> None:
        self.host = host
        self.port = port
        self.stop_event = threading.Event()
        self.server: Optional[socket.socket] = None
        self.accept_thread: Optional[threading.Thread] = None
        self.client_threads: List[threading.Thread] = []
        # Bug #9 fix: Use deque with maxlen to prevent unbounded memory growth
        env_maxlen = os.getenv("GCS_TELEM_MAXLEN")
        maxlen = TELEMETRY_BUFFER_MAXLEN_DEFAULT
        if env_maxlen:
            try:
                candidate = int(env_maxlen)
                if candidate <= 0:
                    raise ValueError
                if candidate < 1000:
                    candidate = 1000
                if candidate > 1_000_000:
                    print(
                        f"[WARN] GCS_TELEM_MAXLEN={candidate} capped at 1000000", file=sys.stderr
                    )
                maxlen = min(candidate, 1_000_000)
            except ValueError:
                print(
                    f"[WARN] invalid GCS_TELEM_MAXLEN={env_maxlen!r}; using default {TELEMETRY_BUFFER_MAXLEN_DEFAULT}",
                    file=sys.stderr,
                )
                maxlen = TELEMETRY_BUFFER_MAXLEN_DEFAULT
        self.samples: deque = deque(maxlen=maxlen)  # ~10MB limit for long tests
        self.lock = threading.Lock()
        self.enabled = True

    def start(self) -> None:
        try:
            addrinfo = socket.getaddrinfo(
                self.host,
                self.port,
                0,
                socket.SOCK_STREAM,
                proto=0,
                flags=socket.AI_PASSIVE if not self.host else 0,
            )
        except socket.gaierror as exc:
            print(f"[WARN] telemetry collector disabled: {exc}", file=sys.stderr)
            self.enabled = False
            return

        last_exc: Optional[Exception] = None
        for family, socktype, proto, _canon, sockaddr in addrinfo:
            try:
                srv = socket.socket(family, socktype, proto)
                try:
                    srv.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                    if family == socket.AF_INET6:
                        srv.setsockopt(socket.IPPROTO_IPV6, socket.IPV6_V6ONLY, 1)
                    srv.bind(sockaddr)
                    srv.listen(8)
                    srv.settimeout(0.5)
                except Exception:
                    srv.close()
                    raise
            except Exception as exc:
                last_exc = exc
                continue

            self.server = srv
            self.accept_thread = threading.Thread(target=self._accept_loop, daemon=True)
            self.accept_thread.start()
            print(f"[{ts()}] telemetry collector listening on {self.host}:{self.port}")
            return

        self.enabled = False
        message = last_exc or RuntimeError("no suitable address family")
        print(f"[WARN] telemetry collector disabled: {message}", file=sys.stderr)
        if self.server:
            try:
                self.server.close()
            except Exception:
                pass
            self.server = None

    def _accept_loop(self) -> None:
        assert self.server is not None
        while not self.stop_event.is_set():
            try:
                conn, addr = self.server.accept()
            except socket.timeout:
                continue
            except OSError:
                break
            except Exception as exc:
                if not self.stop_event.is_set():
                    print(f"[WARN] telemetry accept error: {exc}", file=sys.stderr)
                continue
            thread = threading.Thread(target=self._client_loop, args=(conn, addr), daemon=True)
            thread.start()
            self.client_threads.append(thread)

    def _client_loop(self, conn: socket.socket, addr) -> None:
        peer = f"{addr[0]}:{addr[1]}"
        try:
            conn.settimeout(1.0)
            with conn, conn.makefile("r", encoding="utf-8") as reader:
                for line in reader:
                    if self.stop_event.is_set():
                        break
                    data = line.strip()
                    if not data:
                        continue
                    try:
                        payload = json.loads(data)
                    except json.JSONDecodeError:
                        continue
                    payload.setdefault("collector_ts_ns", time.time_ns())
                    payload.setdefault("source", "drone")
                    payload.setdefault("peer", peer)
                    with self.lock:
                        self.samples.append(payload)
        except Exception:
            # drop connection silently
            pass

    def snapshot(self) -> List[dict]:
        with self.lock:
            # Convert deque to list for compatibility
            return list(self.samples)

    def stop(self) -> None:
        self.stop_event.set()
        if self.server:
            try:
                self.server.close()
            except Exception:
                pass
        if self.accept_thread and self.accept_thread.is_alive():
            self.accept_thread.join(timeout=1.5)
        for thread in self.client_threads:
            if thread.is_alive():
                thread.join(timeout=1.0)

def resolve_under_root(path: Path) -> Path:
    expanded = path.expanduser()
    return expanded if expanded.is_absolute() else ROOT / expanded


def safe_sheet_name(name: str) -> str:
    sanitized = "".join("_" if ch in '[]:*?/\\' else ch for ch in name).strip()
    if not sanitized:
        sanitized = "Sheet"
    return sanitized[:31]


def unique_sheet_name(workbook, base_name: str) -> str:
    base = safe_sheet_name(base_name)
    if base not in workbook.sheetnames:
        return base
    index = 1
    while True:
        suffix = f"_{index}"
        name = base[: 31 - len(suffix)] + suffix
        if name not in workbook.sheetnames:
            return name
        index += 1


def append_dict_sheet(workbook, title: str, rows: List[dict]) -> None:
    if not rows:
        return
    sheet_name = unique_sheet_name(workbook, title)
    ws = workbook.create_sheet(sheet_name)
    headers: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def append_csv_sheet(workbook, path: Path, title: str) -> None:
    if not path.exists():
        return
    rows = None
    for attempt in range(3):
        try:
            with open(path, newline="", encoding="utf-8") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            break
        except OSError as exc:
            if attempt == 2:
                print(f"[WARN] failed to read CSV {path}: {exc}", file=sys.stderr)
            time.sleep(0.1)
        except Exception as exc:
            print(f"[WARN] failed to parse CSV {path}: {exc}", file=sys.stderr)
            return
    if not rows:
        return
    sheet_name = unique_sheet_name(workbook, title)
    ws = workbook.create_sheet(sheet_name)
    for row in rows:
        ws.append(row)


def locate_drone_session_dir(session_id: str) -> Optional[Path]:
    candidates = []
    try:
        candidates.append(resolve_under_root(DRONE_MONITOR_BASE) / session_id)
    except Exception:
        pass
    fallback = Path("/home/dev/research/output/drone") / session_id
    candidates.append(fallback)
    repo_default = ROOT / "output" / "drone" / session_id
    candidates.append(repo_default)
    seen = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        try:
            if candidate.exists():
                return candidate
        except Exception:
            continue
    return None


def export_combined_excel(
    session_id: str,
    summary_rows: List[dict],
    saturation_overview: List[dict],
    saturation_samples: List[dict],
    telemetry_samples: List[dict],
    drone_session_dir: Optional[Path] = None,
    *,
    traffic_mode: str,
    payload_bytes: int,
    event_sample: int,
    min_delay_samples: int,
    pre_gap_s: float,
    duration_s: float,
    inter_gap_s: float,
    sat_search: str,
    sat_delivery_threshold: float,
    sat_loss_threshold_pct: float,
    sat_rtt_spike_factor: float,
) -> Optional[Path]:
    if Workbook is None:
        print("[WARN] openpyxl not available; skipping combined Excel export", file=sys.stderr)
        return None

    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = "run_info"
    info_sheet.append(["generated_utc", ts()])
    info_sheet.append(["session_id", session_id])

    append_dict_sheet(workbook, "gcs_summary", summary_rows)
    append_dict_sheet(workbook, "saturation_overview", saturation_overview)
    append_dict_sheet(workbook, "saturation_samples", saturation_samples)
    append_dict_sheet(workbook, "telemetry_samples", telemetry_samples)

    def _summarize_kinematics(samples: List[dict]) -> List[dict]:
        aggregates: dict[str, dict[str, float]] = {}
        for sample in samples:
            kind = str(sample.get("kind") or "").lower()
            if kind != "kinematics":
                continue
            suite = str(sample.get("suite") or "unknown").strip() or "unknown"
            bucket = aggregates.setdefault(
                suite,
                {
                    "count": 0.0,
                    "pfc_sum": 0.0,
                    "pfc_max": 0.0,
                    "speed_sum": 0.0,
                    "speed_max": 0.0,
                    "altitude_min": float("inf"),
                    "altitude_max": float("-inf"),
                },
            )

            pfc = _as_float(sample.get("predicted_flight_constraint_w"))
            speed = _as_float(sample.get("speed_mps"))
            altitude = _as_float(sample.get("altitude_m"))

            bucket["count"] += 1.0
            if pfc is not None:
                bucket["pfc_sum"] += pfc
                bucket["pfc_max"] = max(bucket["pfc_max"], pfc)
            if speed is not None:
                bucket["speed_sum"] += speed
                bucket["speed_max"] = max(bucket["speed_max"], speed)
            if altitude is not None:
                bucket["altitude_min"] = min(bucket["altitude_min"], altitude)
                bucket["altitude_max"] = max(bucket["altitude_max"], altitude)

        summary_rows: List[dict] = []
        for suite, data in sorted(aggregates.items()):
            count = max(1.0, data["count"])
            altitude_min = "" if math.isinf(data["altitude_min"]) else data["altitude_min"]
            altitude_max = "" if math.isinf(data["altitude_max"]) else data["altitude_max"]
            summary_rows.append(
                {
                    "suite": suite,
                    "samples": int(data["count"]),
                    "pfc_avg_w": _rounded(data["pfc_sum"] / count, 3),
                    "pfc_max_w": _rounded(data["pfc_max"], 3),
                    "speed_avg_mps": _rounded(data["speed_sum"] / count, 3),
                    "speed_max_mps": _rounded(data["speed_max"], 3),
                    "altitude_min_m": _rounded(altitude_min, 3) if altitude_min != "" else "",
                    "altitude_max_m": _rounded(altitude_max, 3) if altitude_max != "" else "",
                }
            )
        return summary_rows

    kinematics_summary = _summarize_kinematics(telemetry_samples)
    append_dict_sheet(workbook, "kinematics_summary", kinematics_summary)

    paper_header = [
        "suite",
        "rekey_ms",
        "blackout_ms",
        "gap_p99_ms",
        "goodput_mbps",
        "loss_pct",
        "rtt_p50_ms",
        "rtt_p95_ms",
        "owd_p50_ms",
        "owd_p95_ms",
        "power_avg_w",
        "power_energy_j",
    ]
    paper_sheet = workbook.create_sheet("paper_tables")
    paper_sheet.append(paper_header)
    ordered_rows: "OrderedDict[str, dict]" = OrderedDict()
    for row in summary_rows:
        suite_name = str(row.get("suite") or "").strip()
        if not suite_name:
            continue
        ordered_rows[suite_name] = row
    paper_rows = list(ordered_rows.items())
    for suite_name, source_row in paper_rows:
        paper_sheet.append([
            suite_name,
            _rounded(source_row.get("rekey_ms"), 3),
            _rounded(source_row.get("blackout_ms"), 3),
            _rounded(source_row.get("gap_p99_ms"), 3),
            _rounded(source_row.get("goodput_mbps"), 3),
            _rounded(source_row.get("loss_pct"), 3),
            _rounded(source_row.get("rtt_p50_ms"), 3),
            _rounded(source_row.get("rtt_p95_ms"), 3),
            _rounded(source_row.get("owd_p50_ms"), 3),
            _rounded(source_row.get("owd_p95_ms"), 3),
            _rounded(source_row.get("power_avg_w"), 6),
            _rounded(source_row.get("power_energy_j"), 6),
        ])

    notes_header = [
        "generated_utc",
        "session_id",
        "traffic_mode",
        "payload_bytes",
        "event_sample",
        "min_delay_samples",
        "pre_gap_s",
        "duration_s",
        "inter_gap_s",
        "sat_search",
        "sat_delivery_threshold",
        "sat_loss_threshold_pct",
        "sat_rtt_spike_factor",
    ]
    notes_sheet = workbook.create_sheet("paper_notes")
    notes_sheet.append(notes_header)
    notes_sheet.append([
        ts(),
        session_id,
        traffic_mode,
        payload_bytes,
        event_sample,
        min_delay_samples,
        round(pre_gap_s, 3),
        round(duration_s, 3),
        round(inter_gap_s, 3),
        sat_search,
        sat_delivery_threshold,
        sat_loss_threshold_pct,
        sat_rtt_spike_factor,
    ])

    if SUMMARY_CSV.exists():
        append_csv_sheet(workbook, SUMMARY_CSV, "gcs_summary_csv")

    if drone_session_dir is None:
        drone_session_dir = locate_drone_session_dir(session_id)
    if drone_session_dir:
        info_sheet.append(["drone_session_dir", str(drone_session_dir)])
        for csv_path in sorted(drone_session_dir.glob("*.csv")):
            append_csv_sheet(workbook, csv_path, csv_path.stem[:31])
    else:
        info_sheet.append(["drone_session_dir", "not_found"])

    if paper_rows and BarChart is not None and Reference is not None:
        row_count = len(paper_rows) + 1
        suite_categories = Reference(paper_sheet, min_col=1, min_row=2, max_row=row_count)

        rekey_chart = BarChart()
        rekey_chart.title = "Rekey vs Blackout (ms)"
        rekey_chart.add_data(
            Reference(paper_sheet, min_col=2, max_col=3, min_row=1, max_row=row_count),
            titles_from_data=True,
        )
        rekey_chart.set_categories(suite_categories)
        rekey_chart.y_axis.title = "Milliseconds"
        rekey_chart.x_axis.title = "Suite"
        paper_sheet.add_chart(rekey_chart, "H2")

        power_chart = BarChart()
        power_chart.title = "Avg Power (W)"
        power_chart.add_data(
            Reference(paper_sheet, min_col=11, max_col=11, min_row=1, max_row=row_count),
            titles_from_data=True,
        )
        power_chart.set_categories(suite_categories)
        power_chart.y_axis.title = "Watts"
        power_chart.x_axis.title = "Suite"
        paper_sheet.add_chart(power_chart, "H18")

    if summary_rows and LineChart is not None and Reference is not None and "gcs_summary" in workbook.sheetnames:
        summary_sheet = workbook["gcs_summary"]
        header_row = next(summary_sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            try:
                pass_col = header_row.index("pass") + 1
                throughput_col = header_row.index("throughput_mbps") + 1
            except ValueError:
                pass_col = throughput_col = None
            if pass_col and throughput_col and len(summary_rows) >= 1:
                chart = LineChart()
                chart.title = "Throughput (Mb/s) vs pass index"
                chart.add_data(
                    Reference(
                        summary_sheet,
                        min_col=throughput_col,
                        min_row=1,
                        max_row=len(summary_rows) + 1,
                    ),
                    titles_from_data=True,
                )
                chart.set_categories(
                    Reference(
                        summary_sheet,
                        min_col=pass_col,
                        min_row=2,
                        max_row=len(summary_rows) + 1,
                    )
                )
                chart.x_axis.title = "Pass"
                chart.y_axis.title = "Throughput (Mb/s)"
                summary_sheet.add_chart(chart, "L2")

    combined_root = resolve_under_root(COMBINED_OUTPUT_DIR)
    combined_dir = combined_root / session_id
    combined_dir.mkdir(parents=True, exist_ok=True)
    info_sheet.append(["gcs_session_dir", str(combined_dir)])
    target_path = combined_dir / f"{session_id}_combined.xlsx"
    for attempt in range(3):
        try:
            buffer = io.BytesIO()
            workbook.save(buffer)
            _atomic_write_bytes(target_path, buffer.getvalue())
            return target_path
        except Exception as exc:  # pragma: no cover - platform specific
            if attempt == 2:
                print(f"[WARN] failed to write combined workbook {target_path}: {exc}", file=sys.stderr)
            time.sleep(0.1)
    return None


def main() -> None:
    log_runtime_environment("gcs_scheduler")
    OUTDIR.mkdir(parents=True, exist_ok=True)
    SUITES_OUTDIR.mkdir(parents=True, exist_ok=True)
    PROXY_STATUS_PATH.parent.mkdir(parents=True, exist_ok=True)
    PROXY_SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)

    auto = AUTO_GCS_CONFIG

    traffic_mode = str(auto.get("traffic") or "blast").lower()
    pre_gap = float(auto.get("pre_gap_s") or 1.0)
    inter_gap = float(auto.get("inter_gap_s") or 15.0)
    duration = float(auto.get("duration_s") or 15.0)
    payload_bytes = int(auto.get("payload_bytes") or 256)
    configured_event_sample = int(auto.get("event_sample") or 100)
    event_sample = max(0, configured_event_sample)
    passes = int(auto.get("passes") or 1)
    rate_pps = int(auto.get("rate_pps") or 0)
    bandwidth_mbps = float(auto.get("bandwidth_mbps") or 0.0)
    constant_rate_defaulted = False
    max_rate_mbps = float(auto.get("max_rate_mbps") or 200.0)
    if traffic_mode == "constant" and bandwidth_mbps <= 0 and rate_pps <= 0:
        bandwidth_mbps = CONSTANT_RATE_MBPS_DEFAULT
        constant_rate_defaulted = True
    if bandwidth_mbps > 0:
        denominator = max(payload_bytes * 8, 1)
        rate_pps = max(1, int((bandwidth_mbps * 1_000_000) / denominator))
    if traffic_mode == "constant" and rate_pps <= 0:
        raise ValueError("AUTO_GCS.rate_pps or bandwidth_mbps must be positive for constant traffic")

    sat_search_cfg = str(auto.get("sat_search") or SATURATION_SEARCH_MODE).lower()
    if sat_search_cfg not in {"auto", "linear", "bisect"}:
        sat_search_cfg = SATURATION_SEARCH_MODE
    sat_delivery_threshold = float(auto.get("sat_delivery_threshold") or SATURATION_DELIVERY_THRESHOLD)
    sat_loss_threshold = float(auto.get("sat_loss_threshold_pct") or SATURATION_LOSS_THRESHOLD)
    sat_spike_factor = float(auto.get("sat_rtt_spike_factor") or SATURATION_RTT_SPIKE)

    min_delay_samples = MIN_DELAY_SAMPLES

    if duration <= 0:
        raise ValueError("AUTO_GCS.duration_s must be positive")
    if pre_gap < 0:
        raise ValueError("AUTO_GCS.pre_gap_s must be >= 0")
    if inter_gap < 0:
        raise ValueError("AUTO_GCS.inter_gap_s must be >= 0")
    if rate_pps < 0:
        raise ValueError("AUTO_GCS.rate_pps must be >= 0")
    if passes <= 0:
        raise ValueError("AUTO_GCS.passes must be >= 1")

    if traffic_mode not in {"blast", "constant", "mavproxy", "saturation"}:
        raise ValueError(f"Unsupported traffic mode: {traffic_mode}")

    constant_target_bandwidth_mbps = 0.0
    if traffic_mode == "constant":
        if bandwidth_mbps > 0:
            constant_target_bandwidth_mbps = bandwidth_mbps
        elif rate_pps > 0:
            constant_target_bandwidth_mbps = (rate_pps * payload_bytes * 8) / 1_000_000
    run_target_bandwidth_mbps = (
        constant_target_bandwidth_mbps if traffic_mode == "constant" else max(0.0, bandwidth_mbps)
    )

    suites_override = auto.get("suites")
    suites = resolve_suites(suites_override)
    if not suites:
        raise RuntimeError("No suites selected for execution")

    session_prefix = str(auto.get("session_prefix") or "session")
    env_session_id = os.environ.get("GCS_SESSION_ID")
    session_id = env_session_id or f"{session_prefix}_{int(time.time())}"
    session_source = "env" if env_session_id else "generated"

    initial_suite = preferred_initial_suite(suites)
    if initial_suite and suites[0] != initial_suite:
        suites = [initial_suite] + [s for s in suites if s != initial_suite]
        print(f"[{ts()}] reordered suites to start with {initial_suite} (from CONFIG)")

    power_capture_enabled = bool(auto.get("power_capture", True))

    telemetry_enabled = bool(auto.get("telemetry_enabled", True))
    telemetry_bind_host = auto.get("telemetry_bind_host") or TELEMETRY_BIND_HOST
    telemetry_port_cfg = auto.get("telemetry_port")
    telemetry_port = TELEMETRY_PORT if telemetry_port_cfg in (None, "") else int(telemetry_port_cfg)

    print(
        f"[{ts()}] traffic={traffic_mode} duration={duration:.1f}s pre_gap={pre_gap:.1f}s "
        f"inter_gap={inter_gap:.1f}s payload={payload_bytes}B event_sample={event_sample} passes={passes} "
        f"rate_pps={rate_pps} sat_search={sat_search_cfg}"
    )
    if traffic_mode == "constant":
        target_msg = f"[{ts()}] constant-rate target {constant_target_bandwidth_mbps:.2f} Mbps (~{rate_pps} pps)"
        if constant_rate_defaulted:
            target_msg += " [default]"
        print(target_msg)
    elif bandwidth_mbps > 0:
        print(f"[{ts()}] bandwidth target {bandwidth_mbps:.2f} Mbps -> approx {rate_pps} pps")
    print(f"[{ts()}] power capture: {'enabled' if power_capture_enabled else 'disabled'}")

    reachable = False
    for attempt in range(8):
        try:
            resp = ctl_send({"cmd": "ping"}, timeout=1.0, retries=1)
            if resp.get("ok"):
                reachable = True
                break
        except Exception:
            pass
        time.sleep(0.5)
    follower_session_id: Optional[str] = None
    if reachable:
        print(f"[{ts()}] follower reachable at {DRONE_HOST}:{CONTROL_PORT}")
        try:
            session_resp = ctl_send({"cmd": "session_info"}, timeout=1.2, retries=2, backoff=0.3)
            if session_resp.get("ok"):
                candidate = str(session_resp.get("session_id") or "").strip()
                if candidate:
                    follower_session_id = candidate
        except Exception as exc:
            print(f"[WARN] session_info fetch failed: {exc}", file=sys.stderr)
    else:
        print(f"[WARN] follower not reachable at {DRONE_HOST}:{CONTROL_PORT}", file=sys.stderr)

    if follower_session_id:
        if env_session_id and follower_session_id != env_session_id:
            print(
                f"[WARN] follower session_id={follower_session_id} disagrees with GCS_SESSION_ID={env_session_id}; using env override",
                file=sys.stderr,
            )
        else:
            session_id = follower_session_id
            session_source = "drone"

    print(f"[{ts()}] session_id={session_id} (source={session_source})")
    os.environ["GCS_SESSION_ID"] = session_id

    drone_session_dir = locate_drone_session_dir(session_id)
    if drone_session_dir:
        print(f"[{ts()}] follower session dir -> {drone_session_dir}")
    else:
        print(f"[WARN] follower session dir missing for session {session_id}", file=sys.stderr)

    session_excel_dir = resolve_under_root(EXCEL_OUTPUT_DIR) / session_id

    offset_ns = 0
    offset_warmup_s = 0.0
    try:
        sync = timesync()
        offset_ns = sync["offset_ns"]
        print(f"[{ts()}] clocks synced: offset_ns={offset_ns} ns, link_rtt~{sync['rtt_ns']} ns")
        if abs(offset_ns) > CLOCK_OFFSET_THRESHOLD_NS:
            offset_warmup_s = 1.0
            print(
                f"[WARN] clock offset {offset_ns / 1_000_000:.1f} ms exceeds {CLOCK_OFFSET_THRESHOLD_NS / 1_000_000:.1f} ms; extending warmup",
                file=sys.stderr,
            )
            print(
                f"[{ts()}] clock skew banner: |offset|={offset_ns / 1_000_000:.1f} ms -> first measurement pass may be noisy",
                flush=True,
            )
    except Exception as exc:
        print(f"[WARN] timesync failed: {exc}", file=sys.stderr)

    telemetry_collector: Optional[TelemetryCollector] = None
    if telemetry_enabled:
        telemetry_collector = TelemetryCollector(telemetry_bind_host, telemetry_port)
        telemetry_collector.start()
        print(f"[{ts()}] telemetry collector -> {telemetry_bind_host}:{telemetry_port}")
    else:
        print(f"[{ts()}] telemetry collector disabled via AUTO_GCS configuration")

    if not bool(auto.get("launch_proxy", True)):
        raise NotImplementedError("AUTO_GCS.launch_proxy=False is not supported")

    gcs_proc: Optional[subprocess.Popen] = None
    log_handle = None
    gcs_proc, log_handle = start_gcs_proxy(suites[0])

    try:
        ready = wait_handshake(timeout=20.0)
        print(f"[{ts()}] initial handshake ready? {ready}")

        summary_rows: List[dict] = []
        saturation_reports: List[dict] = []
        all_rate_samples: List[dict] = []
        telemetry_samples: List[dict] = []

        if traffic_mode == "saturation":
            for idx, suite in enumerate(suites):
                rekey_ms, rekey_mark_ns, rekey_ok_ns = activate_suite(gcs_proc, suite, is_first=(idx == 0))
                outdir = suite_outdir(suite)
                tester = SaturationTester(
                    suite=suite,
                    payload_bytes=payload_bytes,
                    duration_s=duration,
                    event_sample=event_sample,
                    offset_ns=offset_ns,
                    output_dir=outdir,
                    max_rate_mbps=int(max_rate_mbps),
                    search_mode=sat_search_cfg,
                    delivery_threshold=sat_delivery_threshold,
                    loss_threshold=sat_loss_threshold,
                    spike_factor=sat_spike_factor,
                    min_delay_samples=min_delay_samples,
                )
                summary = tester.run()
                summary["rekey_ms"] = rekey_ms
                if rekey_mark_ns is not None:
                    summary["rekey_mark_ns"] = rekey_mark_ns
                if rekey_ok_ns is not None:
                    summary["rekey_ok_ns"] = rekey_ok_ns
                excel_path = tester.export_excel(session_id, session_excel_dir)
                if excel_path:
                    summary["excel_path"] = str(excel_path)
                saturation_reports.append(summary)
                all_rate_samples.extend(dict(record) for record in tester.records)
                if inter_gap > 0 and idx < len(suites) - 1:
                    time.sleep(inter_gap)
            report_path = OUTDIR / f"saturation_summary_{session_id}.json"
            summary_bytes = json.dumps(saturation_reports, indent=2).encode("utf-8")
            try:
                _atomic_write_bytes(report_path, summary_bytes)
                print(f"[{ts()}] saturation summary written to {report_path}")
            except Exception as exc:
                print(f"[WARN] failed to update {report_path}: {exc}", file=sys.stderr)
        else:
            for pass_index in range(passes):
                for idx, suite in enumerate(suites):
                    row = run_suite(
                        gcs_proc,
                        suite,
                        is_first=(pass_index == 0 and idx == 0),
                        duration_s=duration,
                        payload_bytes=payload_bytes,
                        event_sample=event_sample,
                        offset_ns=offset_ns,
                        pass_index=pass_index,
                        traffic_mode=traffic_mode,
                        pre_gap=pre_gap,
                        inter_gap_s=inter_gap,
                        rate_pps=rate_pps,
                        target_bandwidth_mbps=run_target_bandwidth_mbps,
                        power_capture_enabled=power_capture_enabled,
                        clock_offset_warmup_s=offset_warmup_s,
                        min_delay_samples=min_delay_samples,
                        telemetry_collector=telemetry_collector,
                    )
                    summary_rows.append(row)
                    is_last_suite = idx == len(suites) - 1
                    is_last_pass = pass_index == passes - 1
                    if inter_gap > 0 and not (is_last_suite and is_last_pass):
                        time.sleep(inter_gap)

            if summary_rows:
                blackout_records, step_payloads = _enrich_summary_rows(
                    summary_rows,
                    session_id=session_id,
                    drone_session_dir=drone_session_dir,
                    traffic_mode=traffic_mode,
                    pre_gap_s=pre_gap,
                    duration_s=duration,
                    inter_gap_s=inter_gap,
                )
                _append_blackout_records(blackout_records)
                _append_step_results(step_payloads)

            write_summary(summary_rows)

        if telemetry_collector and telemetry_collector.enabled:
            telemetry_samples = telemetry_collector.snapshot()

        if auto.get("export_combined_excel", True):
            combined_path = export_combined_excel(
                session_id=session_id,
                summary_rows=summary_rows,
                saturation_overview=saturation_reports,
                saturation_samples=all_rate_samples,
                telemetry_samples=telemetry_samples,
                drone_session_dir=drone_session_dir,
                traffic_mode=traffic_mode,
                payload_bytes=payload_bytes,
                event_sample=event_sample,
                min_delay_samples=min_delay_samples,
                pre_gap_s=pre_gap,
                duration_s=duration,
                inter_gap_s=inter_gap,
                sat_search=sat_search_cfg,
                sat_delivery_threshold=sat_delivery_threshold,
                sat_loss_threshold_pct=sat_loss_threshold,
                sat_rtt_spike_factor=sat_spike_factor,
            )
            if combined_path:
                print(f"[{ts()}] combined workbook written to {combined_path}")

    finally:
        try:
            ctl_send({"cmd": "stop"})
        except Exception:
            pass

        if gcs_proc and gcs_proc.stdin:
            try:
                gcs_proc.stdin.write("quit\n")
                gcs_proc.stdin.flush()
            except Exception:
                pass
        if gcs_proc:
            try:
                gcs_proc.wait(timeout=5)
            except Exception:
                gcs_proc.kill()

        if log_handle:
            try:
                log_handle.close()
            except Exception:
                pass

        if telemetry_collector:
            telemetry_collector.stop()


if __name__ == "__main__":
    # Test plan:
    # 1. Launch the scheduler with the follower running; verify telemetry collector binds and follower connects.
    # 2. Exercise multiple suites to confirm rekey waits for follower confirmation and no failed rekeys occur.
    # 3. Delete output directories before a run to ensure the scheduler recreates all paths automatically.
    # 4. Stop the telemetry collector briefly and confirm the follower reconnects without aborting the run.
    main()
