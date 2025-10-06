#!/usr/bin/env python3
"""Produce a consolidated text report for a single drone/GCS session.

This tool inspects the combined GCS workbook plus the drone monitor outputs
from the matching session directory. It emits a textual summary covering:

* GCS per-suite throughput / loss / rekey / power metrics
* Telemetry counts gathered by the scheduler
* Drone-side monitoring statistics (CPU, temperature, memory)
* UDP echo loop latency samples
* Power capture summaries recorded on the drone

The report is intended to give operators a single glance at session health
before diving into the raw CSV/JSON artifacts.
"""

from __future__ import annotations

import argparse
import csv
import json
import statistics
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - tooling dependency
    raise SystemExit("openpyxl is required to analyse session workbooks") from exc


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_GCS_OUTPUT = ROOT / "output" / "gcs"
DEFAULT_DRONE_OUTPUT = ROOT / "output" / "drone"


@dataclass
class SuiteMetrics:
    suite: str
    throughput_mbps: float
    loss_pct: float
    rekey_ms: float
    power_avg_w: float
    power_energy_j: float
    power_ok: bool
    rekeys_fail: int


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate consolidated session report")
    parser.add_argument("--session-id", help="Explicit session identifier (e.g. session_1728239927)")
    parser.add_argument(
        "--workbook",
        type=Path,
        help="Path to <session>_combined.xlsx. Defaults to the latest workbook under output/gcs/",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional path to write the report instead of stdout",
    )
    parser.add_argument(
        "--drone-dir",
        type=Path,
        help="Override the drone session directory (defaults to output/drone/<session_id>)",
    )
    return parser.parse_args(argv)


def resolve_path(path: Path) -> Path:
    expanded = path.expanduser()
    return expanded if expanded.is_absolute() else (Path.cwd() / expanded)


def find_latest_workbook(base_dir: Path) -> Optional[Path]:
    if not base_dir.exists():
        return None
    candidates = sorted(
        base_dir.rglob("*_combined.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def load_sheet_records(workbook, sheet_name: str) -> List[Dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: List[Dict[str, object]] = []
    for row in rows[1:]:
        entry: Dict[str, object] = {}
        for header, cell in zip(headers, row):
            if header:
                entry[header] = cell
        if entry:
            records.append(entry)
    return records


def load_run_info(workbook) -> Dict[str, object]:
    if "run_info" not in workbook.sheetnames:
        return {}
    info: Dict[str, object] = {}
    for key, value, *_rest in workbook["run_info"].iter_rows(values_only=True):
        if isinstance(key, str) and key:
            info[key] = value
    return info


def safe_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if value is None:
        return False
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off", ""}:
        return False
    return False


def extract_suite_metrics(rows: Iterable[Dict[str, object]]) -> List[SuiteMetrics]:
    metrics: List[SuiteMetrics] = []
    for row in rows:
        suite = str(row.get("suite") or "unknown")
        metrics.append(
            SuiteMetrics(
                suite=suite,
                throughput_mbps=round(safe_float(row.get("throughput_mbps")), 3),
                loss_pct=round(safe_float(row.get("loss_pct")), 3),
                rekey_ms=round(safe_float(row.get("rekey_ms")), 3),
                power_avg_w=round(safe_float(row.get("power_avg_w")), 3),
                power_energy_j=round(safe_float(row.get("power_energy_j")), 3),
                power_ok=as_bool(row.get("power_capture_ok")),
                rekeys_fail=int(safe_float(row.get("rekeys_fail"))),
            )
        )
    return metrics


def summarise_gcs(metrics: Sequence[SuiteMetrics]) -> Dict[str, object]:
    if not metrics:
        return {}
    best_thr = max(metrics, key=lambda m: m.throughput_mbps)
    worst_loss = max(metrics, key=lambda m: m.loss_pct)
    slowest_rekey = max(metrics, key=lambda m: m.rekey_ms)
    throughput_avg = statistics.mean(m.throughput_mbps for m in metrics)
    loss_avg = statistics.mean(m.loss_pct for m in metrics)
    total_energy = sum(m.power_energy_j for m in metrics)
    power_values = [m.power_avg_w for m in metrics if m.power_avg_w > 0]
    avg_power = statistics.mean(power_values) if power_values else 0.0
    power_gaps = [m.suite for m in metrics if not m.power_ok]
    rekey_failures = [m.suite for m in metrics if m.rekeys_fail > 0]
    return {
        "suite_count": len(metrics),
        "avg_throughput": throughput_avg,
        "avg_loss": loss_avg,
        "best_throughput": (best_thr.suite, best_thr.throughput_mbps),
        "worst_loss": (worst_loss.suite, worst_loss.loss_pct),
        "slowest_rekey": (slowest_rekey.suite, slowest_rekey.rekey_ms),
        "total_energy": total_energy,
        "avg_power": avg_power,
        "power_gaps": power_gaps,
        "rekey_failures": rekey_failures,
    }


def summarise_telemetry(records: Sequence[Dict[str, object]]) -> Dict[str, object]:
    if not records:
        return {"total": 0, "by_kind": {}}
    counts: Dict[str, int] = {}
    timestamps: List[int] = []
    for row in records:
        kind = str(row.get("kind") or "unknown")
        counts[kind] = counts.get(kind, 0) + 1
        ts_value = row.get("timestamp_ns") or row.get("collector_ts_ns")
        try:
            timestamps.append(int(float(ts_value)))
        except (TypeError, ValueError):
            continue
    span = {}
    if timestamps:
        span = {
            "min_ns": min(timestamps),
            "max_ns": max(timestamps),
            "duration_s": (max(timestamps) - min(timestamps)) / 1_000_000_000.0,
        }
    return {"total": len(records), "by_kind": counts, "span": span}


def read_csv(path: Path) -> List[Dict[str, str]]:
    try:
        with path.open("r", encoding="utf-8") as handle:
            return list(csv.DictReader(handle))
    except FileNotFoundError:
        return []
    except Exception as exc:
        print(f"[WARN] failed to read {path}: {exc}", file=sys.stderr)
        return []


def float_field(row: Dict[str, str], key: str) -> Optional[float]:
    value = row.get(key)
    if value is None or value == "":
        return None
    try:
        return float(value)
    except ValueError:
        return None


def summarise_system_monitor(csv_rows: Sequence[Dict[str, str]]) -> Dict[str, object]:
    if not csv_rows:
        return {}
    cpu = [float_field(row, "cpu_percent") for row in csv_rows]
    cpu = [val for val in cpu if val is not None]
    freq = [float_field(row, "cpu_freq_mhz") for row in csv_rows]
    freq = [val for val in freq if val is not None]
    temp = [float_field(row, "cpu_temp_c") for row in csv_rows]
    temp = [val for val in temp if val is not None]
    mem_used = [float_field(row, "mem_used_mb") for row in csv_rows]
    mem_used = [val for val in mem_used if val is not None]
    mem_pct = [float_field(row, "mem_percent") for row in csv_rows]
    mem_pct = [val for val in mem_pct if val is not None]
    result = {"samples": len(csv_rows)}
    if cpu:
        result["cpu_percent"] = stats(cpu)
    if freq:
        result["cpu_freq_mhz"] = stats(freq)
    if temp:
        result["cpu_temp_c"] = stats(temp)
    if mem_used:
        result["mem_used_mb"] = stats(mem_used)
    if mem_pct:
        result["mem_percent"] = stats(mem_pct)
    return result


def stats(values: Sequence[float]) -> Dict[str, float]:
    return {
        "avg": statistics.mean(values),
        "min": min(values),
        "max": max(values),
    }


def summarise_udp_echo(csv_rows: Sequence[Dict[str, str]]) -> Dict[str, object]:
    if not csv_rows:
        return {}
    proc_ns = [float_field(row, "processing_ns") for row in csv_rows]
    proc_ns = [val for val in proc_ns if val is not None]
    if not proc_ns:
        return {"samples": len(csv_rows)}
    avg_ns = statistics.mean(proc_ns)
    return {
        "samples": len(proc_ns),
        "processing_ns": stats(proc_ns),
        "processing_ms_avg": avg_ns / 1_000_000.0,
    }


def summarise_psutil(csv_rows: Sequence[Dict[str, str]]) -> Dict[str, object]:
    if not csv_rows:
        return {}
    cpu: List[float] = []
    rss: List[float] = []
    threads: List[float] = []
    for row in csv_rows:
        cpu_val = float_field(row, "cpu_percent")
        if cpu_val is not None:
            cpu.append(cpu_val)
        rss_val = float_field(row, "rss_bytes")
        if rss_val is not None:
            rss.append(rss_val)
        thread_val = float_field(row, "num_threads")
        if thread_val is not None:
            threads.append(thread_val)
    result: Dict[str, object] = {"samples": len(csv_rows)}
    if cpu:
        result["cpu_percent"] = stats(cpu)
    if rss:
        result["rss_bytes"] = stats(rss)
    if threads:
        result["num_threads"] = stats(threads)
    return result


def summarise_sys_telemetry(csv_rows: Sequence[Dict[str, str]]) -> Dict[str, object]:
    if not csv_rows:
        return {}
    temp: List[float] = []
    freq: List[float] = []
    for row in csv_rows:
        temp_val = float_field(row, "temp_c")
        if temp_val is not None:
            temp.append(temp_val)
        freq_val = float_field(row, "freq_hz")
        if freq_val is not None:
            freq.append(freq_val)
    throttled = {row.get("throttled_hex") for row in csv_rows if row.get("throttled_hex")}
    result: Dict[str, object] = {"samples": len(csv_rows)}
    if temp:
        result["temp_c"] = stats(temp)
    if freq:
        result["freq_hz"] = stats(freq)
    if throttled:
        result["throttled_hex_values"] = sorted(throttled)
    return result


def load_power_summaries(power_dir: Path) -> List[Dict[str, object]]:
    if not power_dir.exists():
        return []
    summaries: List[Dict[str, object]] = []
    for path in sorted(power_dir.glob("*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            data["_source_path"] = path
            summaries.append(data)
        except Exception as exc:
            print(f"[WARN] failed to parse power summary {path}: {exc}", file=sys.stderr)
    return summaries


def format_stats(name: str, payload: Dict[str, object], scale: float = 1.0, unit: str = "") -> str:
    avg = payload.get("avg")
    min_v = payload.get("min")
    max_v = payload.get("max")
    if avg is None or min_v is None or max_v is None:
        return ""
    avg *= scale
    min_v *= scale
    max_v *= scale
    unit_str = f" {unit}" if unit else ""
    return f"{name}: avg={avg:.2f}{unit_str} min={min_v:.2f}{unit_str} max={max_v:.2f}{unit_str}"


def build_report(
    workbook_path: Path,
    run_info: Dict[str, object],
    gcs_metrics: Sequence[SuiteMetrics],
    gcs_summary: Dict[str, object],
    telemetry_summary: Dict[str, object],
    system_monitor: Dict[str, object],
    udp_echo: Dict[str, object],
    psutil_metrics: Dict[str, object],
    sys_telemetry: Dict[str, object],
    power_summaries: Sequence[Dict[str, object]],
) -> str:
    lines: List[str] = []
    lines.append("=== Session Overview ===")
    lines.append(f"Workbook: {workbook_path}")
    session_id = run_info.get("session_id") or run_info.get("Session")
    if session_id:
        lines.append(f"Session ID: {session_id}")
    generated = run_info.get("generated_utc")
    if generated:
        lines.append(f"Generated UTC: {generated}")
    drone_dir = run_info.get("drone_session_dir")
    if drone_dir:
        lines.append(f"Drone session dir: {drone_dir}")
    lines.append("")

    lines.append("=== GCS Summary ===")
    if not gcs_metrics:
        lines.append("No gcs_summary entries present.")
    else:
        header = (
            f"{'Suite':<34} | {'Thr Mb/s':>9} | {'Loss %':>7} | {'Rekey ms':>9} | "
            f"{'Power W':>8} | {'Energy J':>9} | {'Power':>7} | {'RekeyFail':>9}"
        )
        lines.append(header)
        lines.append("-" * len(header))
        for metric in gcs_metrics:
            power_flag = "OK" if metric.power_ok else "MISS"
            lines.append(
                f"{metric.suite:<34} | {metric.throughput_mbps:>9.2f} | {metric.loss_pct:>7.2f} | "
                f"{metric.rekey_ms:>9.2f} | {metric.power_avg_w:>8.3f} | {metric.power_energy_j:>9.3f} | "
                f"{power_flag:>7} | {metric.rekeys_fail:>9}"
            )
        if gcs_summary:
            lines.append("")
            lines.append(f"Suites analysed   : {gcs_summary['suite_count']}")
            lines.append(f"Avg throughput    : {gcs_summary['avg_throughput']:.2f} Mb/s")
            lines.append(f"Avg loss          : {gcs_summary['avg_loss']:.2f} %")
            best_suite, best_thr = gcs_summary["best_throughput"]
            lines.append(f"Best throughput   : {best_suite} @ {best_thr:.2f} Mb/s")
            loss_suite, loss_val = gcs_summary["worst_loss"]
            lines.append(f"Highest loss      : {loss_suite} @ {loss_val:.2f} %")
            rekey_suite, rekey_ms = gcs_summary["slowest_rekey"]
            lines.append(f"Slowest rekey     : {rekey_suite} @ {rekey_ms:.2f} ms")
            lines.append(f"Total energy      : {gcs_summary['total_energy']:.3f} J")
            lines.append(f"Average power     : {gcs_summary['avg_power']:.3f} W")
            if gcs_summary["power_gaps"]:
                lines.append("Missing power data:")
                for suite in gcs_summary["power_gaps"]:
                    lines.append(f"  - {suite}")
            if gcs_summary["rekey_failures"]:
                lines.append("Rekey failures:")
                for suite in gcs_summary["rekey_failures"]:
                    lines.append(f"  - {suite}")
    lines.append("")

    lines.append("=== Telemetry Overview ===")
    total_samples = telemetry_summary.get("total", 0)
    lines.append(f"Total samples: {total_samples}")
    by_kind = telemetry_summary.get("by_kind", {})
    if by_kind:
        for kind, count in sorted(by_kind.items(), key=lambda kv: kv[0]):
            lines.append(f"  {kind}: {count}")
    span = telemetry_summary.get("span") or {}
    if span:
        lines.append(
            f"Timespan: {span['duration_s']:.2f} s (ns {span['min_ns']} … {span['max_ns']})"
        )
    lines.append("")

    lines.append("=== Drone System Monitoring ===")
    if not system_monitor:
        lines.append("system_monitoring CSV not found.")
    else:
        lines.append(f"Samples: {system_monitor['samples']}")
        cpu_stats = system_monitor.get("cpu_percent")
        if cpu_stats:
            lines.append(format_stats("CPU %", cpu_stats))
        freq_stats = system_monitor.get("cpu_freq_mhz")
        if freq_stats:
            lines.append(format_stats("CPU freq", freq_stats, unit="MHz"))
        temp_stats = system_monitor.get("cpu_temp_c")
        if temp_stats:
            lines.append(format_stats("CPU temp", temp_stats, unit="C"))
        mem_used_stats = system_monitor.get("mem_used_mb")
        if mem_used_stats:
            lines.append(format_stats("Mem used", mem_used_stats, unit="MB"))
        mem_pct_stats = system_monitor.get("mem_percent")
        if mem_pct_stats:
            lines.append(format_stats("Mem %", mem_pct_stats))
    lines.append("")

    lines.append("=== UDP Echo Timing ===")
    if not udp_echo:
        lines.append("No UDP echo samples recorded.")
    else:
        lines.append(f"Samples: {udp_echo['samples']}")
        proc_stats = udp_echo.get("processing_ns")
        if proc_stats:
            lines.append(format_stats("Processing", proc_stats, scale=1e-6, unit="ms"))
        avg_ms = udp_echo.get("processing_ms_avg")
        if avg_ms is not None:
            lines.append(f"Average processing delay: {avg_ms:.3f} ms")
    lines.append("")

    lines.append("=== Proxy Process Metrics (psutil) ===")
    if not psutil_metrics:
        lines.append("psutil_proc CSV not found.")
    else:
        lines.append(f"Samples: {psutil_metrics['samples']}")
        if psutil_metrics.get("cpu_percent"):
            lines.append(format_stats("CPU %", psutil_metrics["cpu_percent"]))
        if psutil_metrics.get("rss_bytes"):
            lines.append(format_stats("RSS", psutil_metrics["rss_bytes"], scale=1 / (1024 * 1024), unit="MB"))
        if psutil_metrics.get("num_threads"):
            lines.append(format_stats("Threads", psutil_metrics["num_threads"]))
    lines.append("")

    lines.append("=== Thermal / Clock Telemetry ===")
    if not sys_telemetry:
        lines.append("sys_telemetry CSV not found.")
    else:
        lines.append(f"Samples: {sys_telemetry['samples']}")
        if sys_telemetry.get("temp_c"):
            lines.append(format_stats("Temperature", sys_telemetry["temp_c"], unit="C"))
        if sys_telemetry.get("freq_hz"):
            lines.append(format_stats("Frequency", sys_telemetry["freq_hz"], scale=1e-6, unit="MHz"))
        throttled = sys_telemetry.get("throttled_hex_values")
        if throttled:
            lines.append("Throttled flags observed:")
            for value in throttled:
                lines.append(f"  - {value}")
    lines.append("")

    lines.append("=== Power Captures (Drone) ===")
    if not power_summaries:
        lines.append("No power summaries detected under the session directory.")
    else:
        for summary in power_summaries:
            label = summary.get("label") or summary.get("suite") or "unknown"
            duration = safe_float(summary.get("duration_s"))
            avg_power = safe_float(summary.get("avg_power_w"))
            energy = safe_float(summary.get("energy_j"))
            samples = summary.get("samples", 0)
            csv_path = summary.get("csv_path") or summary.get("_source_path")
            lines.append(
                f"Suite {label}: {duration:.2f}s avg_power={avg_power:.3f}W energy={energy:.3f}J samples={samples} ({csv_path})"
            )
    lines.append("")

    return "\n".join(lines)


def locate_drone_dir(session_id: str, override: Optional[Path], run_info: Dict[str, object]) -> Optional[Path]:
    if override:
        target = resolve_path(override)
        return target if target.exists() else None
    candidates: List[Path] = []
    hinted = run_info.get("drone_session_dir")
    if isinstance(hinted, str) and hinted:
        candidates.append(Path(hinted))
    candidates.append(DEFAULT_DRONE_OUTPUT / session_id)
    # Provide a fallback to ensure we try absolute path from repo root
    candidates.append(ROOT / "output" / "drone" / session_id)
    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.expanduser()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists():
            return resolved
    return None


def main(argv: Optional[Sequence[str]] = None) -> None:
    args = parse_args(argv)

    if args.workbook:
        workbook_path = resolve_path(args.workbook)
        if not workbook_path.exists():
            raise SystemExit(f"Workbook not found: {workbook_path}")
    else:
        workbook_path = find_latest_workbook(DEFAULT_GCS_OUTPUT)
        if not workbook_path:
            raise SystemExit(f"No *_combined.xlsx files found under {DEFAULT_GCS_OUTPUT}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    run_info = load_run_info(workbook)

    session_id = args.session_id or run_info.get("session_id") or run_info.get("Session")
    if not session_id:
        raise SystemExit("Unable to determine session_id; pass --session-id explicitly.")

    gcs_rows = load_sheet_records(workbook, "gcs_summary")
    gcs_metrics = extract_suite_metrics(gcs_rows)
    gcs_summary = summarise_gcs(gcs_metrics)

    telemetry_rows = load_sheet_records(workbook, "telemetry_samples")
    telemetry_summary = summarise_telemetry(telemetry_rows)

    drone_dir = locate_drone_dir(session_id, args.drone_dir, run_info)
    system_monitor_rows: List[Dict[str, str]] = []
    udp_echo_rows: List[Dict[str, str]] = []
    psutil_rows: List[Dict[str, str]] = []
    sys_telemetry_rows: List[Dict[str, str]] = []
    power_summaries: List[Dict[str, object]] = []

    if drone_dir:
        system_monitor_path = drone_dir / f"system_monitoring_{session_id}.csv"
        packet_timing_path = drone_dir / "packet_timing.csv"
        psutil_path = drone_dir / f"psutil_proc_{session_id}.csv"
        sys_telemetry_path = drone_dir / f"sys_telemetry_{session_id}.csv"

        system_monitor_rows = read_csv(system_monitor_path)
        udp_echo_rows = read_csv(packet_timing_path)
        psutil_rows = read_csv(psutil_path)
        sys_telemetry_rows = read_csv(sys_telemetry_path)
        power_summaries = load_power_summaries(drone_dir / "power")
    else:
        print(f"[WARN] drone session directory for {session_id} not found", file=sys.stderr)

    system_monitor_summary = summarise_system_monitor(system_monitor_rows)
    udp_echo_summary = summarise_udp_echo(udp_echo_rows)
    psutil_summary = summarise_psutil(psutil_rows)
    sys_telemetry_summary = summarise_sys_telemetry(sys_telemetry_rows)

    report = build_report(
        workbook_path,
        run_info,
        gcs_metrics,
        gcs_summary,
        telemetry_summary,
        system_monitor_summary,
        udp_echo_summary,
        psutil_summary,
        sys_telemetry_summary,
        power_summaries,
    )

    if args.output:
        output_path = resolve_path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(report, encoding="utf-8")
        print(f"Report written to {output_path}")
    else:
        print(report)


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    main()
